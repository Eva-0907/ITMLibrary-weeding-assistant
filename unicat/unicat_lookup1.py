#!/usr/bin/env python3
"""
UniCat Lookup — ITG Antwerp
Reads weeding_report.xlsx, looks up every WEED item in UniCat,
fetches the number of Belgian libraries holding it, and flips
sole holders (holdings <= 1) from WEED to KEEP.

Usage:
    python3 unicat_lookup.py
    python3 unicat_lookup.py --input weeding_report.xlsx --out weeding_report_unicat.xlsx
    python3 unicat_lookup.py --resume   # skip already-looked-up rows

Requirements:
    pip install openpyxl requests beautifulsoup4

Notes:
- Only processes rows where Recommendation == WEED
- Saves progress every 50 rows so you can resume if interrupted (--resume flag)
- Waits 1-2 seconds between requests to avoid being blocked
- If UniCat is unreachable for a row it marks it "ERROR" and moves on
"""

import time
import random
import argparse
import re
import sys
from pathlib import Path

try:
    import requests
    from bs4 import BeautifulSoup
except ImportError:
    print("Installing required packages...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "requests", "beautifulsoup4", "-q"])
    import requests
    from bs4 import BeautifulSoup

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter


import re
import urllib.parse

def clean_unicat_url(url: str, title: str = "", isbn: str = "") -> str:
    """Rebuild URL using correct UniCat func=search&query= format."""
    if isbn:
        return f"https://www.unicat.be/uniCat?func=search&query={urllib.parse.quote(isbn)}"
    # Clean title: strip brackets and punctuation, take first 4 words
    clean_title = re.sub(r"[\[\];:,!?(){}]", " ", title or "")
    clean_title = re.sub(r"\s+", " ", clean_title).strip()
    q = " ".join(clean_title.split()[:4])
    if not q:
        # Fall back to extracting query from existing URL
        m = re.search(r"query=(.+)$", url or "")
        q = urllib.parse.unquote(m.group(1)) if m else ""
    return f"https://www.unicat.be/uniCat?func=search&query={urllib.parse.quote(q)}"

FILL = {
    "KEEP":   PatternFill("solid", fgColor="DDFFDD"),
    "WEED":   PatternFill("solid", fgColor="FFDDDD"),
    "REVIEW": PatternFill("solid", fgColor="FFFFCC"),
    "SKIP":   PatternFill("solid", fgColor="EEEEEE"),
}

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-GB,en;q=0.9",
}

SESSION = requests.Session()
SESSION.headers.update(HEADERS)
SESSION.verify = False  # UniCat uses a self-signed certificate

# Suppress the SSL warning that would otherwise spam the terminal
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)


# ── UniCat scraper ───────────────────────────────────────────
def fetch_unicat_holdings(url: str, retries: int = 3, debug: bool = False) -> tuple[int | None, str | None]:
    """
    Returns (holdings_count, error_message).
    holdings_count is None if the lookup failed.
    """
    for attempt in range(retries):
        try:
            resp = SESSION.get(url, timeout=10, verify=False)
            resp.raise_for_status()
            raw_html = resp.text
            soup = BeautifulSoup(raw_html, "html.parser")

            if debug:
                print("\n--- RAW HTML (first 5000 chars) ---")
                print(resp.text[:5000])
                print("--- END RAW HTML ---\n")

            # Step 1: extract sysid from search results page
            sysid = None
            # Look for checkAvailability link in the page
            avail_link = soup.find("a", href=re.compile(r"checkAvailability"))
            if not avail_link:
                # Also try finding it in raw HTML
                m = re.search(r"sysid=(\d+)", raw_html)
                if m:
                    sysid = m.group(1)
            else:
                m = re.search(r"sysid=(\d+)", avail_link.get("href", ""))
                if m:
                    sysid = m.group(1)

            if not sysid:
                # No results found at all = 0 holdings
                title_tag = soup.find("title")
                if title_tag and "search" in title_tag.text.lower():
                    return 0, None
                print("\n--- RAW RESPONSE (first 3000 chars) ---")
                print(raw_html[:3000])
                print("--- END RAW RESPONSE ---\n")
                return None, "Could not find sysid in UniCat response"

            # Step 2: fetch availability page
            avail_url = f"https://www.unicat.be/uniCat?func=checkAvailability&sysid={sysid}"
            try:
                avail_resp = SESSION.get(avail_url, timeout=10, verify=False)
                avail_soup = BeautifulSoup(avail_resp.text, "html.parser")
            except requests.exceptions.RequestException:
                return 0, None  # timeout on availability = treat as 0

            # Count library list items in availability response
            # UniCat shows "This title can be found at the following libraries" + bullet list
            lib_items = avail_soup.select("ul li a, ul li")
            if lib_items:
                return len(lib_items), None

            # Check for "not available" message
            not_avail = avail_soup.find(string=re.compile(
                r"(not available|not found|geen|aucune|0 librar)", re.I))
            if not_avail:
                return 0, None

            # Count any links that look like library names
            all_links = avail_soup.find_all("a")
            lib_links = [a for a in all_links if a.text.strip() and
                        not any(x in a.get("href","") for x in ["func=", "javascript", "mailto"])]
            if lib_links:
                return len(lib_links), None

            return 0, None

        except requests.exceptions.RequestException as e:
            if attempt < retries - 1:
                wait = 5 * (attempt + 1)
                print(f"    Retry {attempt+1}/{retries} after {wait}s ({e})")
                time.sleep(wait)
            else:
                return None, str(e)


# ── Column finder ────────────────────────────────────────────
def find_col(ws, header_name: str) -> int | None:
    """Return 1-based column index for a header name, or None."""
    for cell in ws[1]:
        if cell.value and str(cell.value).strip().lower() == header_name.lower():
            return cell.column
    return None


def ensure_col(ws, header_name: str, after_col: int) -> int:
    """Return column index for header, creating it if missing."""
    col = find_col(ws, header_name)
    if col:
        return col
    # Add after the last used column
    new_col = ws.max_column + 1
    cell = ws.cell(row=1, column=new_col, value=header_name)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1E3A5F")
    cell.alignment = Alignment(wrap_text=True)
    return new_col


# ── Main ─────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="UniCat holdings lookup for weeding report")
    parser.add_argument("--input",  default="weeding_report.xlsx", help="Input XLSX (default: weeding_report.xlsx)")
    parser.add_argument("--out",    default="weeding_report_unicat.xlsx", help="Output XLSX (default: weeding_report_unicat.xlsx)")
    parser.add_argument("--resume", action="store_true", help="Skip rows that already have a UniCat holdings value")
    parser.add_argument("--debug", action="store_true", help="Print raw HTML for first result to diagnose parsing issues")
    parser.add_argument("--quiet", action="store_true", help="Only print ISBN lookups and summary, suppress skipped rows")
    args = parser.parse_args()

    input_path = Path(args.input)
    if not input_path.exists():
        print(f"ERROR: {input_path} not found. Run weed.py first.")
        sys.exit(1)

    print(f"Loading {input_path}...")
    wb = openpyxl.load_workbook(input_path)
    ws = wb.active
    print(f"  {ws.max_row - 1:,} data rows")

    # Find required columns
    col_rec      = find_col(ws, "Recommendation")
    col_url      = find_col(ws, "UniCat URL")
    col_title    = find_col(ws, "Title")
    col_reasoning = find_col(ws, "Reasoning")

    if not col_rec or not col_url:
        print("ERROR: Could not find 'Recommendation' or 'UniCat URL' columns.")
        print("Make sure you're using the XLSX produced by weed.py.")
        sys.exit(1)

    # Ensure UniCat columns exist
    col_holdings   = ensure_col(ws, "UniCat Holdings",    ws.max_column)
    col_holders    = ensure_col(ws, "Sole Holder",        ws.max_column)
    col_confidence = ensure_col(ws, "Lookup Confidence",  ws.max_column)

    # Count how many WEED rows need lookup
    weed_rows = []
    for row in range(2, ws.max_row + 1):
        rec = ws.cell(row=row, column=col_rec).value
        if rec == "WEED":
            existing = ws.cell(row=row, column=col_holdings).value
            if args.resume and existing not in (None, "", "ERROR"):
                continue  # already done
            weed_rows.append(row)

    print(f"  {len(weed_rows):,} WEED rows to look up")
    if not weed_rows:
        print("Nothing to do. Use --resume if re-running to skip completed rows.")
        sys.exit(0)

    flipped    = 0
    errors     = 0
    sole_holder = 0
    skipped    = 0
    isbn_rows_looked_up = []

    for i, row in enumerate(weed_rows):
        url   = ws.cell(row=row, column=col_url).value
        title = ws.cell(row=row, column=col_title).value if col_title else f"row {row}"

        title_str = f"  [{i+1}/{len(weed_rows)}] {str(title)[:55]:<55}"

        if not url:
            ws.cell(row=row, column=col_holdings).value   = "No URL"
            ws.cell(row=row, column=col_holders).value    = "No"
            ws.cell(row=row, column=col_confidence).value = "N/A"
            print("→ no URL")
            continue

        # Only look up by ISBN — keyword searches are unreliable
        isbn_col = find_col(ws, "ISBN")
        isbn_val = str(ws.cell(row=row, column=isbn_col).value or "").strip() if isbn_col else ""

        if not isbn_val:
            ws.cell(row=row, column=col_holdings).value   = "No ISBN"
            ws.cell(row=row, column=col_holders).value    = "No"
            ws.cell(row=row, column=col_confidence).value = "N/A"
            if col_reasoning:
                existing = ws.cell(row=row, column=col_reasoning).value or ""
                ws.cell(row=row, column=col_reasoning).value = existing + " | No ISBN — manual UniCat check needed."
            skipped += 1
            time.sleep(0.05)
            continue

        confidence = "HIGH"
        ws.cell(row=row, column=col_confidence).value = confidence
        url = f"https://www.unicat.be/uniCat?func=search&query={urllib.parse.quote(isbn_val)}"
        holdings, err_msg = fetch_unicat_holdings(url, debug=(args.debug and i == 0))
        isbn_rows_looked_up.append(row)

        if holdings is None:
            ws.cell(row=row, column=col_holdings).value = "ERROR"
            ws.cell(row=row, column=col_holders).value  = "?"
            errors += 1
            print(f"\n  Stopped at row {i+1}/{len(weed_rows)}: {err_msg}")
            print(f"  URL attempted: {url}")
            print(f"  Run with --resume to continue once the issue is fixed.")
            wb.save(args.out)
            sys.exit(1)
        else:
            ws.cell(row=row, column=col_holdings).value = holdings
            is_sole = holdings <= 1
            ws.cell(row=row, column=col_holders).value  = "Yes" if is_sole else "No"

            if is_sole:
                # Flip to KEEP
                ws.cell(row=row, column=col_rec).value = "REVIEW"
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = FILL["REVIEW"]

                # Append note to Reasoning
                if col_reasoning:
                    existing_reason = ws.cell(row=row, column=col_reasoning).value or ""
                    ws.cell(row=row, column=col_reasoning).value = (
                        existing_reason + f" | Flagged for REVIEW: sole/no holder in Belgium (UniCat={holdings}, confidence={confidence}) — verify before keeping."
                    )
                sole_holder += 1
                flipped += 1
                pass  # flagged silently
            else:
                pass  # noted silently

        # Print progress summary every 500 rows
        if (i + 1) % 500 == 0:
            isbn_done = (i + 1) - skipped
            print(f"  --- [{i+1}/{len(weed_rows)}] {skipped} skipped (no ISBN) · {isbn_done} looked up · {flipped} flipped to KEEP · {errors} errors ---")

        # Save + print summary every 50 rows
        if (i + 1) % 50 == 0:
            wb.save(args.out)
            isbn_done = (i + 1) - skipped
            print(f"  [{i+1}/{len(weed_rows)}] {skipped} skipped · {isbn_done} looked up · {flipped} flagged for review · {errors} errors — 💾 saved")

        # Polite delay: 1–2 seconds between requests
        time.sleep(1 + random.random())

    # Final save
    wb.save(args.out)

    # Export ISBN-only summary sheet
    summary_path = args.out.replace(".xlsx", "_isbn_summary.xlsx")
    try:
        wb_summary = openpyxl.Workbook()
        ws_summary = wb_summary.active
        ws_summary.title = "ISBN Lookups"
        header_vals = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        ws_summary.append(header_vals)
        for ci in range(1, len(header_vals) + 1):
            cell = ws_summary.cell(row=1, column=ci)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1E3A5F")
        for ri, src_row in enumerate(isbn_rows_looked_up, start=2):
            for ci in range(1, ws.max_column + 1):
                src = ws.cell(row=src_row, column=ci)
                dst = ws_summary.cell(row=ri, column=ci)
                dst.value = src.value
                dst.fill = src.fill
        wb_summary.save(summary_path)
    except Exception as e:
        print(f"  Could not save summary: {e}")

    print(f"\n{'='*55}")
    print(f"Done!")
    print(f"  Skipped (no ISBN):      {skipped:,}")
    print(f"  Looked up (ISBN):       {len(isbn_rows_looked_up):,}")
    print(f"  Flagged for review:     {sole_holder:,}  (sole/no holder in Belgium)")
    print(f"  Errors:                 {errors:,}")
    print(f"  Full output:            {args.out}")
    print(f"  ISBN summary:           {summary_path}")


if __name__ == "__main__":
    main()
