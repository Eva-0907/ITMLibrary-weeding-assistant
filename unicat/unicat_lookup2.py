#!/usr/bin/env python3
"""
UniCat Lookup v2 — ITG Antwerp
Reads UniCat-lookup.xlsx, searches each ISBN in UniCat, and checks
whether any other Belgian library holds a copy.

Logic:
  - Search ISBN in UniCat
  - If the results page contains a "checkAvailability" link → at least one
    other library holds it → leave as WEED
  - If no results / no availability link → no other library holds it →
    flag as REVIEW (may need to keep as sole holder)

Usage:
    python3 unicat_lookup2.py
    python3 unicat_lookup2.py --input data/output/UniCat-lookup.xlsx --out data/output/UniCat-lookup-results.xlsx
    python3 unicat_lookup2.py --resume   # skip already-looked-up rows

Requirements:
    pip install openpyxl requests beautifulsoup4
"""

import time
import random
import argparse
import re
import sys
import urllib.parse
from pathlib import Path

try:
    import requests
    from bs4 import BeautifulSoup
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "requests", "beautifulsoup4", "-q"])
    import requests
    from bs4 import BeautifulSoup

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

FILL = {
    "KEEP":   PatternFill("solid", fgColor="DDFFDD"),
    "WEED":   PatternFill("solid", fgColor="FFDDDD"),
    "REVIEW": PatternFill("solid", fgColor="FFFFCC"),
    "SKIP":   PatternFill("solid", fgColor="EEEEEE"),
}

SESSION = requests.Session()
SESSION.verify = False
SESSION.headers.update({
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-GB,en;q=0.9",
})


# ── UniCat check ────────────────────────────────────────────
def check_unicat_isbn(isbn: str, retries: int = 3, debug: bool = False):
    """
    Search UniCat by ISBN and check if any library holds it.

    Returns:
        ("held",     None)  - checkAvailability link found -> other libraries hold it
        ("not_held", None)  - no results / no availability link -> sole/no holder
        (None,       error) - request failed
    """
    url = f"https://www.unicat.be/uniCat?func=search&query={urllib.parse.quote(isbn)}"

    for attempt in range(retries):
        try:
            resp = SESSION.get(url, timeout=10, verify=False)
            resp.raise_for_status()
            raw_html = resp.text

            if debug:
                print(f"\n--- RAW HTML for ISBN {isbn} (first 5000 chars) ---")
                print(raw_html[:5000])
                print("--- END ---\n")

            soup = BeautifulSoup(raw_html, "html.parser")

            # Key check: is there a checkAvailability link anywhere on the page?
            has_availability = bool(
                soup.find("a", href=re.compile(r"checkAvailability", re.I))
                or re.search(r"checkAvailability", raw_html, re.I)
            )

            if has_availability:
                return "held", None
            else:
                title_tag = soup.find("title")
                page_title = title_tag.text.lower() if title_tag else ""
                if "error" in page_title:
                    return None, f"UniCat error page for ISBN {isbn}"
                return "not_held", None

        except requests.exceptions.Timeout:
            if attempt < retries - 1:
                time.sleep(3 * (attempt + 1))
            else:
                return None, f"Timeout after {retries} attempts"
        except requests.exceptions.RequestException as e:
            if attempt < retries - 1:
                time.sleep(5 * (attempt + 1))
            else:
                return None, str(e)

    return None, "Max retries exceeded"


# ── Column helpers ────────────────────────────────────────────
def find_col(ws, header_name):
    for cell in ws[1]:
        if cell.value and str(cell.value).strip().lower() == header_name.lower():
            return cell.column
    return None


def ensure_col(ws, header_name):
    col = find_col(ws, header_name)
    if col:
        return col
    new_col = ws.max_column + 1
    cell = ws.cell(row=1, column=new_col, value=header_name)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1E3A5F")
    cell.alignment = Alignment(wrap_text=True)
    return new_col


# ── Main ─────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="UniCat ISBN availability check — ITG Antwerp")
    parser.add_argument("--input",  default="data/output/UniCat-lookup.xlsx",        help="Input XLSX (default: data/output/UniCat-lookup.xlsx)")
    parser.add_argument("--out",    default="data/output/UniCat-lookup-results.xlsx", help="Output XLSX (default: data/output/UniCat-lookup-results.xlsx)")
    parser.add_argument("--resume", action="store_true", help="Skip rows already processed")
    parser.add_argument("--debug",  action="store_true", help="Print raw HTML for first lookup")
    args = parser.parse_args()

    input_path = Path(args.input)
    if not input_path.exists():
        print(f"ERROR: {input_path} not found.")
        sys.exit(1)

    print(f"Loading {input_path}...")
    wb = openpyxl.load_workbook(input_path)
    ws = wb.active
    print(f"  {ws.max_row - 1:,} rows")

    col_rec       = find_col(ws, "Recommendation")
    col_isbn      = find_col(ws, "ISBN")
    col_title     = find_col(ws, "Title")
    col_reasoning = find_col(ws, "Reasoning")

    if not col_rec:
        print("ERROR: 'Recommendation' column not found.")
        sys.exit(1)
    if not col_isbn:
        print("ERROR: 'ISBN' column not found.")
        sys.exit(1)

    col_unicat_result = ensure_col(ws, "UniCat Result")
    col_unicat_url    = ensure_col(ws, "UniCat URL")

    rows_to_process = []
    for row in range(2, ws.max_row + 1):
        isbn_val = str(ws.cell(row=row, column=col_isbn).value or "").strip()
        if not isbn_val:
            continue
        if args.resume:
            existing = ws.cell(row=row, column=col_unicat_result).value
            if existing not in (None, "", "ERROR"):
                continue
        rows_to_process.append(row)

    print(f"  {len(rows_to_process):,} rows with ISBN to look up")
    if not rows_to_process:
        print("Nothing to do.")
        sys.exit(0)

    held = 0
    not_held = 0
    errors = 0
    flagged_review = 0

    for i, row in enumerate(rows_to_process):
        isbn_val   = str(ws.cell(row=row, column=col_isbn).value or "").strip()
        title      = str(ws.cell(row=row, column=col_title).value or "") if col_title else f"row {row}"
        unicat_url = f"https://www.unicat.be/uniCat?func=search&query={urllib.parse.quote(isbn_val)}"

        ws.cell(row=row, column=col_unicat_url).value = unicat_url

        result, err_msg = check_unicat_isbn(isbn_val, debug=(args.debug and i == 0))

        if result == "held":
            ws.cell(row=row, column=col_unicat_result).value = "Held elsewhere"
            held += 1

        elif result == "not_held":
            ws.cell(row=row, column=col_unicat_result).value = "Not held elsewhere"
            not_held += 1
            flagged_review += 1
            ws.cell(row=row, column=col_rec).value = "REVIEW"
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = FILL["REVIEW"]
            if col_reasoning:
                existing = ws.cell(row=row, column=col_reasoning).value or ""
                ws.cell(row=row, column=col_reasoning).value = (
                    existing + f" | REVIEW: no other Belgian library holds this (ISBN={isbn_val}) — verify before weeding."
                )

        else:
            ws.cell(row=row, column=col_unicat_result).value = "ERROR"
            errors += 1
            print(f"\n  [{i+1}/{len(rows_to_process)}] ERROR: {err_msg}")
            print(f"  Title: {title[:60]}")
            print(f"  ISBN:  {isbn_val}")
            print(f"  URL:   {unicat_url}")
            print(f"  Run with --resume to continue once the issue is fixed.")
            wb.save(args.out)
            sys.exit(1)

        if (i + 1) % 50 == 0:
            wb.save(args.out)
            print(f"  [{i+1}/{len(rows_to_process)}] held={held} · not_held={not_held} · review={flagged_review} — 💾 saved")

        time.sleep(1 + random.random())

    wb.save(args.out)

    print(f"\n{'='*55}")
    print(f"Done!")
    print(f"  Looked up:          {len(rows_to_process):,}")
    print(f"  Held elsewhere:     {held:,}  -> stays WEED")
    print(f"  Not held elsewhere: {not_held:,}  -> flagged REVIEW")
    print(f"  Errors:             {errors:,}")
    print(f"  Output saved:       {args.out}")


if __name__ == "__main__":
    main()
