#!/usr/bin/env python3
"""
unicat_apply_cache.py — ITG Antwerp Library Weeding Project

Offline companion to unicat_lookup3.py. Applies previously-cached UniCat
sole-holder lookup results (unicat_cache.json, keyed by ISBN) to a fresh
weeding report, WITHOUT re-querying unicat.be over the network.

Replicates the exact logic of unicat_lookup3.py's main():
  - "held"     -> "UniCat Result" = "Held elsewhere", recommendation unchanged
  - "not_held" -> "UniCat Result" = "Not held elsewhere"
       * current recommendation == WEED  -> upgraded to REVIEW, row re-filled,
         Reasoning gets the same "REVIEW: no other Belgian library holds this
         (ISBN=...) — verify before weeding." suffix
       * else (KEEP/REVIEW/SKIP)          -> recommendation unchanged,
         Reasoning gets "Note: ITG is the sole Belgian holder (ISBN=...)."
  - ISBN not found in cache -> "UniCat Result" = "Not in cache" (flagged so it
    can be looked up live later); recommendation unchanged.

Only operates on the active sheet (Library Collection), matching
unicat_lookup3.py's behaviour.

Usage:
    python3 unicat_apply_cache.py
    python3 unicat_apply_cache.py --input data/output/weeding_report9.xlsx \
        --cache data/cache/unicat_cache.json --out data/output/UniCat-lookup-results3.xlsx
"""

import sys
import json
import argparse
import urllib.parse
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "--break-system-packages", "-q"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

FILL = {
    "KEEP":   PatternFill("solid", fgColor="DDFFDD"),
    "WEED":   PatternFill("solid", fgColor="FFDDDD"),
    "REVIEW": PatternFill("solid", fgColor="FFFFCC"),
    "SKIP":   PatternFill("solid", fgColor="EEEEEE"),
}


def find_col(ws, header_name):
    for cell in ws[1]:
        if cell.value and str(cell.value).strip().lower() == header_name.lower():
            return cell.column
    return None


def ensure_col(ws, header_name):
    col = find_col(ws, header_name)
    if col:
        return col
    new_col = ws.max_column + 1
    cell = ws.cell(row=1, column=new_col, value=header_name)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1E3A5F")
    cell.alignment = Alignment(wrap_text=True)
    return new_col


def main():
    parser = argparse.ArgumentParser(description="Apply cached UniCat ISBN results (offline) — ITG Antwerp")
    parser.add_argument("--input", default="data/output/weeding_report9.xlsx", help="Input XLSX (weed script output)")
    parser.add_argument("--cache", default="data/cache/unicat_cache.json", help="Cached UniCat results, keyed by ISBN")
    parser.add_argument("--out",   default="data/output/UniCat-lookup-results3.xlsx", help="Output XLSX")
    args = parser.parse_args()

    input_path = Path(args.input)
    cache_path = Path(args.cache)
    if not input_path.exists():
        print(f"ERROR: {input_path} not found."); sys.exit(1)
    if not cache_path.exists():
        print(f"ERROR: {cache_path} not found."); sys.exit(1)

    print(f"Loading {input_path}...")
    wb = openpyxl.load_workbook(input_path)
    ws = wb.active
    print(f"  {ws.max_row - 1:,} rows")

    print(f"Loading cache {cache_path}...")
    cache = json.load(open(cache_path, encoding="utf-8"))
    print(f"  {len(cache):,} cached ISBN results")

    col_rec       = find_col(ws, "Recommendation")
    col_isbn      = find_col(ws, "ISBN")
    col_reasoning = find_col(ws, "Reasoning")

    if not col_rec:
        print("ERROR: 'Recommendation' column not found."); sys.exit(1)
    if not col_isbn:
        print("ERROR: 'ISBN' column not found."); sys.exit(1)

    col_unicat_result = ensure_col(ws, "UniCat Result")
    col_unicat_url    = ensure_col(ws, "UniCat URL")

    held = 0
    not_held = 0
    flagged_review = 0
    not_in_cache = 0
    total_isbn_rows = 0

    for row in range(2, ws.max_row + 1):
        isbn_val = str(ws.cell(row=row, column=col_isbn).value or "").strip()
        if not isbn_val:
            continue
        total_isbn_rows += 1

        unicat_url = f"https://www.unicat.be/uniCat?func=search&query={urllib.parse.quote(isbn_val)}"
        ws.cell(row=row, column=col_unicat_url).value = unicat_url

        entry = cache.get(isbn_val)
        if not entry:
            ws.cell(row=row, column=col_unicat_result).value = "Not in cache"
            not_in_cache += 1
            continue

        result = entry.get("result")

        if result == "held":
            ws.cell(row=row, column=col_unicat_result).value = "Held elsewhere"
            held += 1

        elif result == "not_held":
            ws.cell(row=row, column=col_unicat_result).value = "Not held elsewhere"
            not_held += 1
            current_rec = str(ws.cell(row=row, column=col_rec).value or "").strip()

            if current_rec == "WEED":
                ws.cell(row=row, column=col_rec).value = "REVIEW"
                flagged_review += 1
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = FILL["REVIEW"]
                if col_reasoning:
                    existing = ws.cell(row=row, column=col_reasoning).value or ""
                    ws.cell(row=row, column=col_reasoning).value = (
                        existing + f" | REVIEW: no other Belgian library holds this (ISBN={isbn_val}) — verify before weeding."
                    )
            else:
                if col_reasoning:
                    existing = ws.cell(row=row, column=col_reasoning).value or ""
                    ws.cell(row=row, column=col_reasoning).value = (
                        existing + f" | Note: ITG is the sole Belgian holder (ISBN={isbn_val})."
                    )
        else:
            ws.cell(row=row, column=col_unicat_result).value = "Not in cache"
            not_in_cache += 1

    wb.save(args.out)

    print(f"\n{'='*55}")
    print("Done! (offline — reused cached results, no live UniCat queries)")
    print(f"  ISBN rows total:    {total_isbn_rows:,}")
    print(f"  Held elsewhere:     {held:,}  -> recommendation unchanged")
    print(f"  Sole holder (ITG):  {not_held:,}  -> of which {flagged_review:,} WEED rows upgraded to REVIEW")
    print(f"                         (KEEP/REVIEW/SKIP rows kept their recommendation; sole-holder noted in Reasoning)")
    print(f"  Not in cache:       {not_in_cache:,}  (would need a live lookup)")
    print(f"  Output saved:       {args.out}")


if __name__ == "__main__":
    main()
