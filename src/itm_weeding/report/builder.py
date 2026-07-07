"""Excel report generation."""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from itm_weeding.core.helpers import barnard_label, gf


class ExcelReporter:
    """Generate the Excel workbook used for the final weeding report."""
    
    def __init__(self):
        """Initialize the worksheet layout, formatting, and column configuration."""
        self.headers = [
            "Title", "Author", "Year", "Type", "Bib#", "ISBN",
            "Barnard", "Retention Flag", "Call Number", "Language", "Location",
            "Recommendation", "Circulated", "Historically Significant",
            "Historical Reasons", "Triggered Rules", "Check UniCat", "Reasoning",
        ]
        self.col_widths = [50, 25, 6, 8, 10, 14, 30, 8, 12, 6, 14, 14, 10, 12, 40, 60, 50, 60]
        self.fill_map = {"WEED": "FFDDDD", "KEEP": "DDFFDD", "REVIEW": "FFFFCC", "SKIP": "EEEEEE"}
        self.header_fill = PatternFill("solid", fgColor="1E3A5F")
        self.header_fill_dept = PatternFill("solid", fgColor="4A235A")
    
    def _format_row(self, row):
        """Convert one processed record into the values written to the Excel sheet."""
        result = row["result"]
        flags = result["flags"]
        rec = row["rec"]
        circulated = any(f["criterion"] == "Circulation" for f in flags)
        triggered = "; ".join(f"{f['criterion']}: {f['detail']}" for f in flags)
        rec_val = result["recommendation"]
        
        # Format UniCat check result
        unicat_check = ""
        if rec_val == "WEED" and row.get("isbn"):
            unicat_result = row.get("unicat_result")
            if unicat_result == "held":
                unicat_check = "HELD (Belgium)"
            elif unicat_result == "not_held":
                unicat_check = "Not held"
            else:
                unicat_check = "CHECK"
        
        return rec_val, [
            row["title"],
            row["author"],
            row["year"],
            row["rec_type"],
            gf(rec, "ID"),
            row["isbn"],
            barnard_label(gf(rec, "U4")),
            result["retention"] or "",
            gf(rec, "U5"),
            gf(rec, "U3"),
            row["location"],
            rec_val,
            "Yes" if circulated else "No",
            "Yes" if result["historically_significant"] else "No",
            "; ".join(result["historical_reasons"]),
            triggered,
            unicat_check,
            result["reasoning"],
        ]
    
    def _write_sheet(self, ws, sheet_rows, hdr_fill):
        """Write a set of processed rows to a worksheet with formatting applied."""
        # Header
        for ci, h in enumerate(self.headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = hdr_fill
            cell.alignment = Alignment(wrap_text=True)
        
        # Data
        for excel_row, row in enumerate(sheet_rows, 2):
            rec_val, values = self._format_row(row)
            for ci, v in enumerate(values, 1):
                cell = ws.cell(row=excel_row, column=ci, value=v)
                cell.fill = PatternFill("solid", fgColor=self.fill_map.get(rec_val, "FFFFFF"))
                cell.alignment = Alignment(wrap_text=True)
        
        # Widths / freeze / filter
        for ci, w in enumerate(self.col_widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.freeze_panes = "A2"
        if ws.max_row > 1:
            ws.auto_filter.ref = ws.dimensions
    
    def export(self, rows, out_path):
        """Export the processed rows to an Excel workbook with library and department sheets."""
        # Split rows into library vs department
        lib_rows = [r for r in rows if "dep-a" not in r["location"].lower()]
        dept_rows = [r for r in rows if "dep-a" in r["location"].lower()]
        
        # Create workbook
        wb = Workbook()
        ws_lib = wb.active
        ws_lib.title = "Library Collection"
        self._write_sheet(ws_lib, lib_rows, self.header_fill)
        
        ws_dept = wb.create_sheet("Department Books")
        self._write_sheet(ws_dept, dept_rows, self.header_fill_dept)
        
        print(f"  Sheet 'Library Collection': {len(lib_rows):,} records")
        print(f"  Sheet 'Department Books':   {len(dept_rows):,} records")
        
        wb.save(out_path)


def export_xlsx(rows, out_path):
    """Convenience wrapper for exporting processed rows to an Excel workbook."""
    reporter = ExcelReporter()
    reporter.export(rows, out_path)
