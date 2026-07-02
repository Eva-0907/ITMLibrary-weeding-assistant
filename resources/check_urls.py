#!/usr/bin/env python3
"""
URL Checker — ITG Antwerp Weeding Project
Reads the RIS file, extracts all digital links (L2 and UR fields),
checks whether each URL is still active, and produces an Excel report.

Usage:
    python3 check_urls.py books_1950-1990_book_infile.txt
    python3 check_urls.py books_1950-1990_book_infile.txt --out url_check_results.xlsx
    python3 check_urls.py books_1950-1990_book_infile.txt --concurrency 5
    python3 check_urls.py books_1950-1990_book_infile.txt --resume

Results:
    OK       — URL returned HTTP 200 (accessible)
    BROKEN   — URL returned 4xx/5xx or could not connect
    REDIRECT — URL redirected to another location
    TIMEOUT  — No response within timeout
"""

import re, sys, time, threading, argparse
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed

try:
    import requests
    import urllib3
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "requests", "-q"])
    import requests, urllib3
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

SESSION = requests.Session()
SESSION.headers.update({"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"})

STATUS_FILLS = {
    "OK":       PatternFill("solid", fgColor="DDFFDD"),
    "BROKEN":   PatternFill("solid", fgColor="FFDDDD"),
    "REDIRECT": PatternFill("solid", fgColor="FFFFCC"),
    "TIMEOUT":  PatternFill("solid", fgColor="FFE5CC"),
    "PENDING":  PatternFill("solid", fgColor="F5F5F5"),
}

def parse_ris(path):
    for enc in ("utf-8-sig", "utf-8", "cp850", "windows-1252"):
        try:
            text = Path(path).read_text(encoding=enc)
            if "\ufffd" not in text:
                print(f"  Encoding: {enc}")
                return text
        except (UnicodeDecodeError, LookupError):
            continue
    return Path(path).read_text(encoding="cp850", errors="replace")

def parse_records(text):
    records = []
    for block in re.split(r"\nER\s*-", text):
        rec = {}
        for line in block.strip().splitlines():
            m = re.match(r"^([A-Z][A-Z0-9])\s*-\s*(.*)$", line)
            if m:
                tag, val = m.group(1), m.group(2).strip()
                if tag in rec:
                    if isinstance(rec[tag], list): rec[tag].append(val)
                    else: rec[tag] = [rec[tag], val]
                else:
                    rec[tag] = val
        if rec.get("T1") or rec.get("TI"):
            records.append(rec)
    return records

def gf(rec, *tags):
    for t in tags:
        v = rec.get(t)
        if v: return v[0] if isinstance(v, list) else v
    return ""

def extract_urls(rec):
    urls = []
    for tag in ("L2", "UR"):
        v = rec.get(tag, "")
        vals = [v] if isinstance(v, str) else v
        for val in vals:
            val = re.sub(r"[\x00-\x1f]", "", val).strip()
            if val.startswith("http"):
                urls.append((tag, val))
    return urls

def check_url(url, timeout=10):
    """Returns (status, http_code, final_url)"""
    try:
        resp = SESSION.head(url, timeout=timeout, verify=False, allow_redirects=False)
        if resp.status_code in (405, 501):
            resp = SESSION.get(url, timeout=timeout, verify=False, allow_redirects=False, stream=True)
            resp.close()
        if resp.status_code in (301, 302, 303, 307, 308):
            return "REDIRECT", resp.status_code, resp.headers.get("Location", "")
        elif resp.status_code == 200:
            return "OK", 200, url
        else:
            return "BROKEN", resp.status_code, url
    except requests.exceptions.Timeout:
        return "TIMEOUT", 0, url
    except Exception:
        return "BROKEN", 0, url

def main():
    parser = argparse.ArgumentParser(description="URL checker — ITG weeding project")
    parser.add_argument("ris", help="Path to RIS file")
    parser.add_argument("--out", default="url_check_results.xlsx")
    parser.add_argument("--concurrency", type=int, default=5)
    parser.add_argument("--resume", action="store_true")
    args = parser.parse_args()

    print(f"Reading {args.ris}...")
    text = parse_ris(args.ris)
    records = parse_records(text)
    print(f"  Parsed {len(records):,} records")

    # Extract all URL rows
    url_rows = []
    for rec in records:
        urls = extract_urls(rec)
        if not urls: continue
        bib   = gf(rec, "ID")
        title = gf(rec, "T1", "TI")
        year  = re.search(r"\d{4}", gf(rec, "Y1", "PY", "DA") or "")
        year  = year.group() if year else ""
        barn  = gf(rec, "U4")
        for tag, url in urls:
            url_rows.append({"bib": bib, "title": title, "year": year,
                             "barnard": barn, "tag": tag, "url": url})

    print(f"  {len(url_rows):,} URLs across {sum(1 for r in records if extract_urls(r)):,} records")

    # Resume: load already-checked URLs
    already_done = {}  # url -> (status, code, final)
    if args.resume and Path(args.out).exists():
        wb_old = openpyxl.load_workbook(args.out, data_only=True)
        ws_old = wb_old.active
        hdrs_old = [ws_old.cell(row=1, column=c).value for c in range(1, ws_old.max_column+1)]
        if "URL" in hdrs_old and "Status" in hdrs_old:
            uc = hdrs_old.index("URL") + 1
            sc = hdrs_old.index("Status") + 1
            hc = hdrs_old.index("HTTP Code") + 1
            fc = hdrs_old.index("Redirect To") + 1
            for row in range(2, ws_old.max_row+1):
                url = ws_old.cell(row=row, column=uc).value
                sta = ws_old.cell(row=row, column=sc).value
                if url and sta and sta not in ("PENDING", None):
                    already_done[url] = (sta,
                        ws_old.cell(row=row, column=hc).value,
                        ws_old.cell(row=row, column=fc).value)
        print(f"  Resuming — {len(already_done)} already checked, skipping")

    # Build workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "URL Check Results"
    hfill = PatternFill("solid", fgColor="1F3864")
    hdrs = ["Bib#", "Title", "Year", "Barnard", "Field", "URL", "Status", "HTTP Code", "Redirect To"]
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = hfill
        c.alignment = Alignment(wrap_text=True)

    # Write all rows
    for ri, row in enumerate(url_rows, 2):
        if row["url"] in already_done:
            sta, code, final = already_done[row["url"]]
        else:
            sta, code, final = "PENDING", "", ""
        vals = [row["bib"], row["title"], row["year"], row["barnard"],
                row["tag"], row["url"], sta, code or "", final or ""]
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill = STATUS_FILLS.get(sta, STATUS_FILLS["PENDING"])

    for ci, w in enumerate([8, 55, 6, 18, 6, 65, 10, 10, 60], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(args.out)

    # Check pending URLs
    to_check = [(i, row) for i, row in enumerate(url_rows) if row["url"] not in already_done]
    print(f"\nChecking {len(to_check):,} URLs ({args.concurrency} parallel)...\n")

    lock = threading.Lock()
    counts = {"OK": 0, "BROKEN": 0, "REDIRECT": 0, "TIMEOUT": 0}
    done_n = [0]

    def check_row(item):
        i, row = item
        return i, *check_url(row["url"])

    with ThreadPoolExecutor(max_workers=args.concurrency) as ex:
        futures = {ex.submit(check_row, item): item for item in to_check}
        for future in as_completed(futures):
            i, status, code, final = future.result()
            with lock:
                er = i + 2
                ws.cell(row=er, column=7, value=status)
                ws.cell(row=er, column=8, value=code if code else "")
                ws.cell(row=er, column=9, value=final if final != url_rows[i]["url"] else "")
                rf = STATUS_FILLS.get(status, STATUS_FILLS["PENDING"])
                for col in range(1, len(hdrs)+1):
                    ws.cell(row=er, column=col).fill = rf
                counts[status] = counts.get(status, 0) + 1
                done_n[0] += 1
                if done_n[0] % 50 == 0 or done_n[0] == len(to_check):
                    wb.save(args.out)
                    print(f"  [{done_n[0]:>4}/{len(to_check)}]  "
                          f"OK={counts['OK']}  BROKEN={counts['BROKEN']}  "
                          f"REDIRECT={counts['REDIRECT']}  TIMEOUT={counts['TIMEOUT']}  💾")

    wb.save(args.out)
    print(f"\n{'='*50}")
    print(f"Done! Output: {args.out}")
    print(f"  OK:       {counts['OK']:>5}")
    print(f"  BROKEN:   {counts['BROKEN']:>5}")
    print(f"  REDIRECT: {counts['REDIRECT']:>5}")
    print(f"  TIMEOUT:  {counts['TIMEOUT']:>5}")

if __name__ == "__main__":
    main()
