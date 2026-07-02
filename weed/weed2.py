#!/usr/bin/env python3
"""
Library Weeding Agent — ITG Antwerp
Tropical Medicine & Global Health collection

Usage:
    python weed.py collection.ris
    python weed.py collection.ris --students students.csv --staff staff.csv
    python weed.py collection.ris --students students.csv --staff staff.csv --out results.xlsx

Requirements:
    pip install openpyxl

The script is fully local — no API calls, no internet required.
All rules match the browser artifact exactly.
"""

import re
import sys
import csv
import argparse
from pathlib import Path
from datetime import datetime

# ── Historical titles ────────────────────────────────────────
HISTORICAL_TITLES = [
    "on the origin of species","micrographia","an introduction to the study of experimental medicine",
    "the double helix","molecular biology of the gene","snow on cholera",
    "report on the sanitary condition of the labouring population of great britain",
    "modern epidemiology","epidemiology an introduction","epidemiology in medicine",
    "the ghost map","the germ theory of disease","an inquiry into the causes and effects of the variolae vaccinae",
    "tropical diseases a practical guide","mansons tropical diseases","manson's tropical diseases",
    "the prevention of malaria","principles and practice of medicine","grays anatomy","gray's anatomy",
    "the great influenza","and the band played on","tropical medicine and hygiene",
    "clinical parasitology","bacteriology and immunity","the bacteria","principles of bacteriology",
    "medical microbiology","the pathogenesis of viral infections","medical virology","fields virology",
    "human parasitology","foundations of parasitology","medical parasitology",
    "public health administration","oxford textbook of public health",
    "harrisons principles of internal medicine","cecil textbook of medicine",
    "atlas of human anatomy","robbins pathologic basis of disease",
    "a history of medicine","the greatest benefit to mankind",
    "janeway's immunobiology","the immune system","basic immunology",
    "cellular and molecular immunology","global health 101","foundations of tropical medicine",
    "manual of tropical medicine","parasitic diseases","essential malariology",
    "atlas of human parasitology","medical helminthology","medical entomology for students",
    "vectors of human disease","mosquitoes of the world","biology of disease vectors",
    "emerging infectious diseases","mpox and other poxviruses",
    "principles of biomedical ethics","ethics and professionalism in medicine",
    "medical ethics a very short introduction","spillover","oxford handbook of clinical medicine",
    "clinical examination","molecular pathology","diagnostic pathology",
]

OUTBREAK_TIMELINE = [
    {"event": "First smallpox vaccination (Jenner)",         "keywords": ["smallpox","vaccine","vaccin","jenner","variol"]},
    {"event": "Broad Street cholera outbreak (John Snow)",   "keywords": ["cholera","snow","broad street","sanit"]},
    {"event": "Discovery of TB bacteria (Koch)",             "keywords": ["tuberculosis","tb","koch","mycobacterium"]},
    {"event": "Mosquito transmission of Malaria (Ross)",     "keywords": ["malaria","plasmodium","ross","mosquito","anopheles"]},
    {"event": "Yellow fever transmission (Walter Reed)",     "keywords": ["yellow fever","walter reed","aedes","flavivirus"]},
    {"event": "Discovery of Chagas disease",                 "keywords": ["chagas","trypanosoma","carlos chagas"]},
    {"event": "1918 Influenza pandemic",                     "keywords": ["influenza","spanish flu","1918","pandemic"]},
    {"event": "Malaria eradication campaigns",               "keywords": ["malaria","eradication","ddt"]},
    {"event": "Global smallpox eradication programme",       "keywords": ["smallpox","eradication","who","vaccine"]},
    {"event": "HIV/AIDS",                                    "keywords": ["hiv","aids","immunodeficiency","retrovirus"]},
    {"event": "Global expansion of Dengue",                  "keywords": ["dengue","aedes","arbovirus"]},
    {"event": "Mpox outbreaks",                              "keywords": ["mpox","monkeypox","poxvirus"]},
    {"event": "Ebola outbreak in West Africa",               "keywords": ["ebola","filovirus","hemorrhagic","west africa"]},
    {"event": "COVID-19 pandemic",                           "keywords": ["covid","sars-cov","coronavirus"]},
]

BARNARD = {
    "A":"Reference","AA":"Encyclopedias","AB":"Scientific and medical dictionaries","AC":"Directories","AR":"Education, Teaching",
    "B":"Natural Science","BA":"Scientific literature","BB":"Statistics","BC":"Chemistry","BCB":"Biochemistry","BCC":"Clinical chemistry",
    "BD":"Nutrition","BI":"Microscopy","BJ":"Biology","BJC":"Molecular biology","BJS":"Evolution","BJT":"Genetics","BJX":"Ecology",
    "BK":"Botany","BL":"Zoology","BM":"Anatomy","BN":"Histology","BQ":"Physiology",
    "C":"Internal Medicine","CBA":"Medical education","CBE":"Medical ethics","CBN":"General practice","CBP":"Group practice","CBX":"Nursing","CZ":"Traditional medicine",
    "D":"History of Medicine",
    "E":"Epidemiology","EA":"Disease ecology","EB":"Epidemiology","EC":"Disease surveillance","EH":"Demography","EK":"Mortality","EL":"Mortality","EP":"Medical geography","EV":"Human ecology",
    "G":"Toxicology","GK":"Food poisoning","GL":"Venomous animals",
    "H":"Immunology, Infectious diseases","HD":"Immunology","HG":"Immunodiagnosis","HI":"Vaccination","HW":"Infectious diseases",
    "I":"Mycology",
    "J":"Bacteriology","JC":"Tuberculosis","JD":"Leprosy","JI":"Gonorrhea","JJ":"Bacterial meningitis","JK":"Cholera","JM":"Tetanus","JN":"Plague","JP":"Brucellosis","JS":"Salmonelloses","JX":"Treponematoses","JYC":"Lyme disease","JZ":"Leptospiroses",
    "K":"Virology","KA":"Rickettsioses","KG":"Chlamydia","KGJ":"Trachoma","KI":"Smallpox","KK":"Rabies","KP":"Arboviral diseases","KPA":"Yellow fever","KPD":"Dengue","KPH":"Hemorrhagic fever","KQ":"Encephalitis","KR":"Retroviruses","KRC":"HIV, AIDS","KS":"Hepatitis","KT":"Poliomyelitis",
    "L":"Parasitology","LA":"Protozoology","LC":"Babesiasis","LD":"Coccidiasis","LDC":"Cryptosporidiasis","LDP":"Pneumocystosis","LE":"Toxoplasmosis","LF":"Malaria","LL":"Amebiasis","LN":"Sleeping sickness","LNX":"Animal trypanosomiasis","LP":"Chagas disease","LQ":"Leishmaniasis","LV":"Giardiasis",
    "M":"Helminthology","MH":"Schistosomiasis","MK":"Echinococcosis","MQ":"Ascariasis","MQS":"Strongyloidiasis","MR":"Ankylostomiasis","MS":"Filariases","MU":"Onchocerciasis","MV":"Trichinelliasis",
    "N":"Entomology","NB":"Ectoparasites","NC":"Vector-borne diseases","ND":"Vector control","NDI":"Insecticides","NDY":"Biological control","NG":"Mites","NGT":"Ticks","NK":"Bugs","NL":"Flies","NO":"Mosquitoes","NP":"Simuliidae","NT":"Tsetse flies",
    "O":"Intermediate Hosts",
    "P":"Pathology","PM":"Cancer","PY":"Hematology","PYB":"Blood groups","PYE":"Blood transfusion","PYH":"Anemia","PYHR":"Hemoglobinopathies","PYHS":"Sickle cell anemia","PYK":"Nutritional anemias",
    "Q":"Diagnosis","QD":"Clinical medicine","QH":"Laboratory medicine","QM":"Laboratory medicine","QR":"Radiology",
    "R":"Pharmacology","RJ":"Pharmacology","RN":"Therapeutics","RNA":"Clinical trials","RNR":"Drug resistance","RS":"Drug supply",
    "S":"Public Health","SA":"Environmental hygiene","SD":"Water hygiene","SJA":"Disaster relief","SO":"Public health","SOA":"Health organizations","SOB":"Health information","SOC":"Health evaluation","SOCA":"Health research","SOD":"Health economics","SOE":"Health financing","SOF":"Health manpower","SOG":"Hospitals","SOH":"Health education","SON":"Medical sociology","SOP":"Migrants health","SP":"Preventive medicine","SPA":"Disinfection","SPZ":"Hospital infections","SS":"Travel health","ST":"Urban health","SW":"Naval health",
    "U":"Medical Specialties","UB":"Tropical medicine","UBR":"Tropical medicine","UH":"Cardiology","UI":"Neurology","UJ":"Psychiatry","UK":"Ophthalmology","UL":"Oto-rhino-laryngology","UO":"Respiratory diseases","UP":"Gastroenterology","UPT":"Enteric infections","UQ":"Endocrinology","UR":"Dermatology","US":"Urology","UT":"Sexology","UTH":"Fertility","UTT":"Family planning","UTU":"Sexually transmitted diseases","UV":"Gynecology","UW":"Obstetrics","UWB":"Midwives","UWE":"Mother and child health","UX":"Pediatrics","UY":"Geriatrics",
    "V":"Surgery","VG":"Anesthesia","W":"Oral Health",
    "X":"Veterinary Sciences","XC":"Veterinary medicine","XE":"Epizootiology","XH":"Immunology","XHW":"Infectious diseases","XJ":"Microbiology","XK":"Virology","XL":"Parasitology","XM":"Helminthology","XN":"Entomology","XO":"Zoonoses","XS":"Hygiene","XW":"Animal husbandry","XXB":"Cattle","XXN":"Small ruminants","XXO":"Pigs","XXU":"Poultry","XXZ":"Wild animals","XZ":"Laboratory animals",
    "Y":"Agriculture","YA":"Agriculture","YC":"Field crops","YH":"Pest control",
    "Z":"Geography, Sociology","ZB":"Geography","ZVM":"Development","ZVO":"Economics","ZVP":"Political Science","ZVU":"Management","ZVY":"Colonial administration","ZY":"Population issues","ZYG":"Population movement",
}

RETENTION_FLAGS = {
    "A":"H3","AA":"H3","AB":"H3","AC":"H3","AR":"H2",
    "BA":"H2","BB":"H2","BC":"H3","BCB":"H3","BCC":"H3","BD":"H3","BE":"H3","BF":"H3",
    "BJ":"H2","BJC":"H2","BJS":"H2","BJT":"H2","BJX":"H2","BK":"H2","BL":"H2","BM":"H1","BN":"H2","BQ":"H2",
    "C":"H2","CBA":"H2","CBE":"H1","CBN":"H3","CBP":"H3","CBX":"H3","CZ":"H2",
    "D":"H1",
    "EA":"H2","EB":"H1","EC":"H2","EH":"H2","EK":"H2","EL":"H2","EP":"H2","EQ":"H2","ES":"H2","ET":"H2","EV":"H2",
    "G":"H2","GK":"H3","GL":"H2",
    "H":"H2","HD":"H2","HG":"H2","HI":"H2","HW":"H2",
    "I":"H2",
    "J":"H2","JC":"H2","JD":"H2","JI":"H2","JJ":"H2","JK":"H2","JM":"H2","JN":"H2","JP":"H2","JS":"H2","JX":"H2","JYC":"H2","JZ":"H2",
    "K":"H2","KA":"H2","KG":"H2","KGJ":"H2","KI":"H2","KK":"H2","KP":"H2","KPA":"H2","KPD":"H2","KPH":"H2","KQ":"H2","KR":"H2","KRC":"H2","KS":"H2","KT":"H2",
    "L":"H2","LA":"H2","LC":"H2","LD":"H2","LDC":"H2","LDP":"H2","LE":"H2","LF":"H1","LL":"H2","LN":"H2","LNX":"H2","LP":"H2","LQ":"H2","LV":"H2",
    "M":"H2","MH":"H2","MK":"H2","MQ":"H2","MQS":"H2","MR":"H2","MS":"H2","MU":"H2","MV":"H2","MW":"H2",
    "N":"H2","NB":"H2","NC":"H2","ND":"H2","NDI":"H3","NDY":"H2","NG":"H2","NGT":"H2","NK":"H2","NL":"H2","NO":"H2","NP":"H2","NT":"H2",
    "O":"H2",
    "P":"H2","PM":"H2","PY":"H2","PYB":"H2","PYE":"H2","PYH":"H2","PYHR":"H2","PYHS":"H2","PYK":"H2",
    "QD":"H3","QH":"H3","QM":"H3","QR":"H3",
    "RJ":"H2","RN":"H3","RNA":"H2","RNR":"H3","RS":"H3",
    "S":"H2","SA":"H2","SB":"H2","SC":"H2","SD":"H2","SJA":"H2","SO":"H2","SOA":"H1","SOB":"H2","SOC":"H2","SOCA":"H2","SOD":"H2","SOE":"H2","SOF":"H2","SOG":"H2","SOH":"H2","SON":"H2","SOP":"H2","SP":"H2","SPA":"H2","SPZ":"H2","SS":"H2","ST":"H2","SW":"H2",
    "U":"H2","UB":"H2","UBR":"H2","UH":"H2","UI":"H2","UJ":"H2","UK":"H2","UL":"H2","UO":"H2","UP":"H2","UPT":"H2","UQ":"H2","UR":"H2","US":"H2","UT":"H2","UTH":"H2","UTT":"H2","UTU":"H2","UV":"H2","UW":"H2","UWB":"H2","UWE":"H2","UX":"H2","UY":"H2",
    "V":"H2","VG":"H3","W":"H3",
    "X":"H2","XC":"H2","XE":"H2","XH":"H2","XHW":"H2","XJ":"H2","XK":"H2","XL":"H2","XM":"H2","XN":"H2","XO":"H2","XS":"H2","XW":"H3","XXB":"H3","XXN":"H3","XXO":"H3","XXU":"H3","XXZ":"H2","XZ":"H2",
    "Y":"H2","YA":"H2","YC":"H2","YH":"H3",
    "Z":"H2","ZB":"H2","ZI":"H1","ZR":"H1","ZV":"H2","ZVL":"H2","ZVM":"H2","ZVN":"H2","ZVO":"H2","ZVP":"H2","ZVU":"H2","ZVY":"H2","ZY":"H2","ZYG":"H2",
}

CONGO_TERMS   = ["congo","belgian congo","zaire","belgique","belgium","kinshasa","leopoldville","brazzaville","katanga"]
IRCB_TERMS    = ["ircb","institut royal colonial belge","royal belgian colonial institute",
                 "koninklijk belgisch koloniaal instituut","arsc","academie royale des sciences coloniales",
                 "royal academy of colonial sciences","koninklijke academie voor koloniale wetenschappen"]


# ── Helpers ──────────────────────────────────────────────────
def gf(rec, *tags):
    for t in tags:
        v = rec.get(t)
        if v:
            return v[0] if isinstance(v, list) else v
    return ""

def barnard_label(code):
    if not code:
        return ""
    c = code.strip().upper()
    if c in BARNARD:
        return f"{c} — {BARNARD[c]}"
    for i in range(len(c)-1, 0, -1):
        p = c[:i]
        if p in BARNARD:
            return f"{c} — {BARNARD[p]}"
    return c

def get_retention_flag(code):
    if not code:
        return None
    c = code.strip().upper()
    if c in RETENTION_FLAGS:
        return RETENTION_FLAGS[c]
    for i in range(len(c)-1, 0, -1):
        p = c[:i]
        if p in RETENTION_FLAGS:
            return RETENTION_FLAGS[p]
    return None

def get_year(rec):
    raw = gf(rec, "Y1", "PY", "DA")
    m = re.search(r"\d{4}", raw)
    return int(m.group()) if m else None

def get_isbn(rec):
    return re.sub(r"[^0-9X]", "", gf(rec, "SN"), flags=re.IGNORECASE)

def get_authors(rec):
    all_authors = []
    for tag in ("A1", "A2"):
        v = rec.get(tag, [])
        all_authors += v if isinstance(v, list) else [v]
    if not all_authors:
        return ""
    if len(all_authors) <= 2:
        return "; ".join(all_authors)
    return "; ".join(all_authors[:2]) + " et al."

def make_circ_key(barnard, call_num, year):
    b = (barnard or "").strip().upper()
    c = re.sub(r"/.*$", "", (call_num or "").strip())
    c = re.sub(r"M$", "", c, flags=re.IGNORECASE).strip()
    y = (year or "").strip()
    return f"{b}|{c}|{y}"

def get_circ_key(rec):
    b = gf(rec, "U4").strip().upper()
    c = re.sub(r"M$", "", gf(rec, "U5").strip(), flags=re.IGNORECASE).strip()
    raw = gf(rec, "Y1", "PY", "DA")
    m = re.search(r"\d{4}", raw)
    y = m.group() if m else ""
    return make_circ_key(b, c, y)

def unicat_url(rec):
    import urllib.parse
    isbn = get_isbn(rec)
    if isbn:
        return f"https://www.unicat.be/uniCat?func=search&query={urllib.parse.quote(isbn)}"
    # Clean title: strip punctuation, brackets, take first 4 meaningful words
    raw_title = gf(rec, "T1", "TI")
    clean_title = re.sub(r"[\[\];:,!?(){}]", " ", raw_title)
    clean_title = re.sub(r"\s+", " ", clean_title).strip()
    title  = " ".join(clean_title.split()[:4])
    author = " ".join(gf(rec, "A1", "A2").split()[:2])
    q = " ".join(filter(None, [title, author]))
    return f"https://www.unicat.be/uniCat?func=search&query={urllib.parse.quote(q)}"

def is_historical_title(title):
    t = re.sub(r"[^a-z0-9 ]", " ", title.lower())
    t = re.sub(r"\s+", " ", t).strip()
    for ht in HISTORICAL_TITLES:
        if ht in t or (len(t) >= 20 and ht.startswith(t[:20])):
            return True
    return False

def matches_outbreak(title, abstract, keywords):
    hay = (title + " " + abstract + " " + keywords).lower()
    return [ev["event"] for ev in OUTBREAK_TIMELINE
            if any(k in hay for k in ev["keywords"])]


# ── RIS parser ───────────────────────────────────────────────
def parse_ris(text):
    records = []
    for block in re.split(r"\nER\s*-", text):
        rec = {}
        for line in block.strip().splitlines():
            m = re.match(r"^([A-Z][A-Z0-9])\s*-\s*(.*)$", line)
            if not m:
                continue
            tag, val = m.group(1), m.group(2).strip()
            if tag in rec:
                if isinstance(rec[tag], list):
                    rec[tag].append(val)
                else:
                    rec[tag] = [rec[tag], val]
            else:
                rec[tag] = val
        if rec.get("T1") or rec.get("TI"):
            records.append(rec)
    return records


# ── Circulation loader ───────────────────────────────────────
def load_circulation(path, delimiter, b_col=1, c_col=2, y_col=3):
    borrowed = set()
    try:
        with open(path, encoding="utf-8-sig", errors="replace") as f:
            for line in f:
                cols = line.rstrip("\n").split(delimiter)
                if len(cols) <= max(b_col, c_col, y_col):
                    continue
                barnard  = cols[b_col].strip()
                call_num = cols[c_col].strip()
                year     = cols[y_col].strip()
                if barnard and call_num and re.search(r"\d", call_num):
                    borrowed.add(make_circ_key(barnard, call_num, year))
    except FileNotFoundError:
        print(f"  Warning: file not found: {path}", file=sys.stderr)
    return borrowed


# ── Rule engine ──────────────────────────────────────────────
def apply_rules(rec, all_records, borrowed_bibs, isbn_counts):
    year          = get_year(rec)
    isbn          = get_isbn(rec)
    rec_type      = gf(rec, "TY").upper()
    title         = gf(rec, "T1", "TI")
    abstract      = gf(rec, "N2")
    kw_list       = rec.get("KW", [])
    keywords      = ", ".join(kw_list if isinstance(kw_list, list) else [kw_list])
    publisher     = gf(rec, "PB").lower()
    note          = gf(rec, "N1")
    doc_type      = gf(rec, "U1")
    e_status      = gf(rec, "U2")
    barnard       = gf(rec, "U4")
    retention     = get_retention_flag(barnard)
    has_e_version = ("e-also" in e_status.lower() or "-h" in doc_type.lower() or bool(gf(rec, "L2")))

    flags         = []
    keep_override = False
    hard_rule     = None

    def flag(criterion, detail, severity):
        flags.append({"criterion": criterion, "detail": detail, "severity": severity})

    # ── E-only skip ──
    if "e-only" in e_status.lower() or "-e" in doc_type.lower():
        flag("E-only", "No physical copy", "keep")
        return {"recommendation": "SKIP", "keep_override": False, "hard_rule": "SKIP",
                "flags": flags, "retention": retention, "historically_significant": False,
                "historical_reasons": [], "reasoning": "Skipped — e-only."}

    # ── H1 class ──
    if retention == "H1":
        flag("Retention flag", f"H1 — Core historical field ({barnard_label(barnard)}) — retain indefinitely", "keep")
        keep_override = True

    # ── Dissertation ──
    if (rec_type in ("THES","THESIS","DISS")
            or "dissertation" in doc_type.lower()
            or "dissertation" in title.lower()):
        flag("Document type", "Doctoral dissertation — always keep", "keep")
        keep_override = True

    # ── Proceedings ──
    is_proceedings = (rec_type in ("CONF","CPAPER")
                      or "proceedings" in title.lower()
                      or "conference" in title.lower())
    if is_proceedings:
        if not has_e_version:
            flag("Document type", "Conference proceedings (no e-version) — keep", "keep")
            keep_override = True
        else:
            flag("Document type", "Conference proceedings — e-version exists", "weed")
            hard_rule = "WEED"

    # ── Liber amicorum ──
    if "liber amicorum" in title.lower() or "festschrift" in title.lower():
        flag("Document type", "Liber amicorum / Festschrift — always keep", "keep")
        keep_override = True

    # ── Congo / Belgium ──
    hay = (title + " " + abstract + " " + publisher + " " + keywords).lower()
    if any(t in hay for t in CONGO_TERMS):
        flag("Regional relevance", "Congo / Belgium — always keep", "keep")
        keep_override = True

    # ── IRCB / ARSC ──
    hay2 = (title + " " + abstract + " " + publisher + " " + keywords + " " + note).lower()
    if any(t in hay2 for t in IRCB_TERMS):
        flag("IRCB/ARSC", "IRCB or ARSC publication — always keep", "keep")
        keep_override = True

    # ── Historical title list ──
    if is_historical_title(title):
        flag("Historical title", "Listed in Top Historical Titles for Tropical Medicine", "keep")
        keep_override = True

    # ── Outbreak relevance ──
    outbreaks = matches_outbreak(title, abstract, keywords)
    if outbreaks:
        flag("Outbreak relevance", "Relates to: " + "; ".join(outbreaks[:3]), "keep")
        keep_override = True

    # ── Circulation ──
    if borrowed_bibs and get_circ_key(rec) in borrowed_bibs:
        flag("Circulation", "Borrowed at least once (2019–2026) — keep", "keep")
        keep_override = True

    # ── Barnard (informational) ──
    if barnard:
        flag("Barnard", barnard_label(barnard) + (f" [{retention}]" if retention else ""), "keep")

    # ── Weed rules (only if not already kept) ──
    if not keep_override:
        # Duplicate
        if isbn and isbn_counts.get(isbn, 0) > 1:
            flag("Duplicate", f"{isbn_counts[isbn]} copies — this is a duplicate", "weed")
            hard_rule = "WEED"

        if year is not None:
            if 1950 <= year <= 1990:
                if retention == "H3":
                    flag("Publication date", f"Published {year} [H3 class] — auto-weed", "weed")
                    hard_rule = "WEED"
                elif retention == "H2":
                    flag("Publication date", f"Published {year} [H2 class] — auto-weed (no special flags)", "weed")
                    hard_rule = "WEED"
                else:
                    flag("Publication date", f"Published {year} (1950–1990)", "weed")
                    hard_rule = "WEED"
            elif year < 1950:
                flag("Publication date", f"Published {year} (pre-1950) — potential historical value", "keep")
                keep_override = True
            elif year > 1990:
                flag("Publication date", f"Published {year} (post-1990, not circulated) — auto-weed", "weed")
                hard_rule = "WEED"
        else:
            flag("Publication date", "No publication year — flagged for manual review", "review")
            hard_rule = "REVIEW"

    recommendation = "KEEP" if keep_override else (hard_rule or "REVIEW")
    hist_criteria  = {"Historical title","Outbreak relevance","Regional relevance","IRCB/ARSC","Retention flag"}
    historically   = keep_override and any(f["criterion"] in hist_criteria for f in flags)
    hist_reasons   = [f["detail"] for f in flags if f["severity"] == "keep"]
    reasoning      = ". ".join(f["detail"] for f in flags)

    return {
        "recommendation":        recommendation,
        "keep_override":         keep_override,
        "hard_rule":             hard_rule,
        "flags":                 flags,
        "retention":             retention,
        "historically_significant": historically,
        "historical_reasons":    hist_reasons,
        "reasoning":             reasoning,
    }


# ── XLSX export ──────────────────────────────────────────────
def export_xlsx(rows, out_path):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl not found — installing...", file=sys.stderr)
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Weeding Report"

    headers = [
        "Title","Author","Year","Type","Bib#","ISBN",
        "Barnard","Retention Flag","Call Number","Language",
        "Recommendation","Circulated","Historically Significant",
        "Historical Reasons","Triggered Rules","UniCat URL","Reasoning",
    ]

    # Header row
    fill_map = {"WEED": "FFDDDD", "KEEP": "DDFFDD", "REVIEW": "FFFFCC", "SKIP": "EEEEEE"}
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True)

    for row in rows:
        result  = row["result"]
        flags   = result["flags"]
        rec     = row["rec"]
        ri      = row["row_index"]

        circulated = any(f["criterion"] == "Circulation" for f in flags)
        triggered  = "; ".join(f"{f['criterion']}: {f['detail']}" for f in flags)
        rec_val    = result["recommendation"]

        values = [
            row["title"], row["author"], row["year"], row["rec_type"],
            gf(rec, "ID"), row["isbn"],
            barnard_label(gf(rec, "U4")), result["retention"] or "",
            gf(rec, "U5"), gf(rec, "U3"),
            rec_val,
            "Yes" if circulated else "No",
            "Yes" if result["historically_significant"] else "No",
            "; ".join(result["historical_reasons"]),
            triggered,
            unicat_url(rec),
            result["reasoning"],
        ]

        ws.append(values)
        row_fill = PatternFill("solid", fgColor=fill_map.get(rec_val, "FFFFFF"))
        for ci in range(1, len(headers)+1):
            cell = ws.cell(row=ri+2, column=ci)
            cell.fill = row_fill
            cell.alignment = Alignment(wrap_text=True)

    # Column widths
    col_widths = [50,25,6,8,10,14,30,8,12,6,14,10,12,40,60,50,60]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(out_path)


# ── Main ─────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Library Weeding Agent — ITG Antwerp")
    parser.add_argument("ris",               help="Path to .ris file")
    parser.add_argument("--students",        help="Student loans CSV (semicolon-delimited)")
    parser.add_argument("--staff",           help="Staff loans CSV/TSV (tab-delimited)")
    parser.add_argument("--out",             default="weeding_report2.xlsx", help="Output XLSX filename")
    args = parser.parse_args()

    # Load RIS
    print(f"Reading {args.ris}...")
    # Try encodings in order — latin-1 covers all byte values so it never fails
    ris_text = None
    for enc in ("utf-8-sig", "utf-8", "cp850", "windows-1252"):
        try:
            ris_text = Path(args.ris).read_text(encoding=enc)
            # Sanity check: if we see replacement chars, try next encoding
            if "�" not in ris_text:
                print(f"  Encoding detected: {enc}")
                break
            else:
                ris_text = None
        except (UnicodeDecodeError, LookupError):
            continue
    if ris_text is None:
        # Last resort: cp850 handles legacy DOS/library system exports
        ris_text = Path(args.ris).read_text(encoding="cp850", errors="replace")
        print("  Encoding detected: cp850 (fallback)")
    records  = parse_ris(ris_text)
    print(f"  Parsed {len(records):,} records")

    # Load circulation data
    borrowed = set()
    if args.students:
        print(f"Loading student loans: {args.students}")
        s = load_circulation(args.students, delimiter=";")
        borrowed |= s
        print(f"  {len(s):,} student loan records")
    if args.staff:
        print(f"Loading staff loans: {args.staff}")
        s = load_circulation(args.staff, delimiter="\t")
        borrowed |= s
        print(f"  {len(s):,} staff loan records")
    if borrowed:
        print(f"  {len(borrowed):,} unique circulated items total")

    # Pre-compute ISBN counts for duplicate detection
    isbn_counts = {}
    for rec in records:
        isbn = get_isbn(rec)
        if isbn:
            isbn_counts[isbn] = isbn_counts.get(isbn, 0) + 1

    # Process
    print(f"\nProcessing {len(records):,} records...")
    t0      = datetime.now()
    output_rows = []
    counts  = {"KEEP": 0, "WEED": 0, "REVIEW": 0, "SKIP": 0}

    for i, rec in enumerate(records):
        result = apply_rules(rec, records, borrowed, isbn_counts)
        rec_val = result["recommendation"]
        counts[rec_val] = counts.get(rec_val, 0) + 1

        output_rows.append({
            "row_index": i,
            "rec":       rec,
            "result":    result,
            "title":     gf(rec, "T1", "TI") or "Untitled",
            "author":    get_authors(rec),
            "year":      get_year(rec) or "",
            "rec_type":  gf(rec, "TY").upper(),
            "isbn":      get_isbn(rec),
        })

        if (i+1) % 500 == 0 or (i+1) == len(records):
            elapsed = (datetime.now() - t0).total_seconds()
            rate    = (i+1) / elapsed if elapsed > 0 else 0
            print(f"  {i+1:>6,} / {len(records):,}  "
                  f"KEEP={counts['KEEP']}  WEED={counts['WEED']}  "
                  f"REVIEW={counts['REVIEW']}  SKIP={counts['SKIP']}  "
                  f"({rate:.0f}/s)")

    elapsed = (datetime.now() - t0).total_seconds()
    print(f"\nDone in {elapsed:.1f}s  ({len(records)/elapsed:.0f} records/sec)")
    print(f"  KEEP:   {counts['KEEP']:>6,}")
    print(f"  WEED:   {counts['WEED']:>6,}")
    print(f"  REVIEW: {counts['REVIEW']:>6,}")
    print(f"  SKIP:   {counts['SKIP']:>6,}")

    # Export
    print(f"\nExporting to {args.out}...")
    export_xlsx(output_rows, args.out)
    print(f"Done — {args.out}")


if __name__ == "__main__":
    main()
