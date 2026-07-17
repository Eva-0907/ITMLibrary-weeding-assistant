#!/usr/bin/env python3
"""
Library Weeding Agent — ITG Antwerp
Tropical Medicine & Global Health collection

Usage:
    python weed.py data/books_2_ris.txt
    python weed.py data/books_2_ris.txt --students data/Uitleen_2019-2026.csv --staff data/Uitleen_collega's.csv
    python weed.py data/books_2_ris.txt --students data/Uitleen_2019-2026.csv --staff data/Uitleen_collega's.csv --out data/output/weeding_report9.xlsx

Changes in v8 vs v7:
  - OUTBREAK_TIMELINE keywords tightened to multi-word phrases and named-entity anchors
    (e.g. "1918 influenza", "trypanosoma cruzi", "smallpox eradication") rather than
    broad single disease-name terms that matched too widely.
  - matches_outbreak() now searches title + abstract only; RIS keywords field excluded
    as they are auto-assigned from abstracts and unreliable.

Changes in v9 vs v8:
  - Each OUTBREAK_TIMELINE event now has a "barnard_prefixes" gate. A keyword match
    only counts as a true outbreak match if the record's own Barnard classification
    starts with one of the listed prefixes for that event (e.g. cholera requires
    Barnard "JK", malaria requires "LF"/"NO"/"N"/"ND"/"NC"). This prevents cross-domain
    false positives where an incidental mention in the abstract (e.g. "Koch" appearing
    in an unrelated pediatrics or botany textbook) triggered a keep.
  - With the Barnard gate in place, several keywords could be safely broadened back to
    single disease names (e.g. "cholera", "malaria", "smallpox", "yellow fever",
    "tuberculosis", "influenza", "dengue") since the gate restricts false cross-domain
    matches while still catching genuinely on-topic historical titles that the v8
    multi-word-phrase requirement was missing (e.g. "The cholera problem",
    "Smallpox" by Dixon, "Yellow fever" by Strode).
  - Ambiguous surname-only keywords removed (e.g. standalone "koch", "ross") in favour
    of disease names + the Barnard gate, which is more robust than relying on a
    historical figure's surname alone.
  - Fixed translation-duplicate logic: the English copy is the retained edition. A
    non-English copy of the same work is now always weeded when an English edition
    exists in the collection — even if the non-English copy independently qualifies
    for a keep rule (H1 class, outbreak relevance, etc.). Previously the code did the
    opposite: an independently-protected non-English copy would either keep its
    protection (an earlier, incorrect attempt at this fix) or be downgraded to REVIEW
    rather than weeded outright. The WHO/FAO final downgrade check is also skipped for
    translation duplicates, since the WEED decision here is final — the WHO/FAO
    content is already covered by the retained English edition. The English edition
    itself is never affected, since it does not carry the translation_duplicate flag.

Requirements:
    pip install openpyxl

The script is fully local — no API calls, no internet required.
All rules match the browser artifact exactly.
"""

import re
import sys
import csv
import argparse
import unicodedata
from pathlib import Path
from datetime import datetime

# ── Historical titles ─────────────────────────────────────────────────────────
# Each entry is (title_fragment, author_fragments_or_None).
# - title_fragment: matched case-insensitively (punctuation stripped) against the
#   record title; must also cover ≥60% of the record title's words.
# - author_fragments: space-separated surnames/tokens — ANY ONE must appear in
#   the record's author field (A1/A2). None = title is unique enough on its own.
# Source: ITG canonical list with verified author(s).
HISTORICAL_TITLES = [
    # ── Natural science & general science ──────────────────────────────────────
    ("on the origin of species",                                         "darwin"),
    ("micrographia",                                                     "hooke"),
    ("an introduction to the study of experimental medicine",            "bernard"),
    ("the double helix",                                                 "watson"),
    ("molecular biology of the gene",                                    "watson"),

    # ── History of medicine ────────────────────────────────────────────────────
    ("a history of medicine",                                            "porter"),
    ("the greatest benefit to mankind",                                  "porter"),

    # ── Epidemiology & public health ───────────────────────────────────────────
    ("snow on cholera",                                                  "snow"),
    ("report on the sanitary condition of the labouring population",     "chadwick"),
    ("modern epidemiology",                                              "rothman"),
    ("epidemiology an introduction",                                     "rothman"),
    ("epidemiology in medicine",                                         "hennekens buring"),
    ("public health administration",                                     "winslow"),
    ("oxford textbook of public health",                                 "detels"),

    # ── Tropical medicine ──────────────────────────────────────────────────────
    ("mansons tropical diseases",                                        "manson cook"),   # orig. + editors
    ("manson s tropical diseases",                                       "manson cook"),
    ("tropical diseases a practical guide",                              "cook zumla"),
    ("the prevention of malaria",                                        "ross"),
    ("tropical medicine and hygiene",                                    None),            # various eds., title unique enough
    ("foundations of tropical medicine",                                 "jawetz"),
    ("manual of tropical medicine",                                      "hunter swartzwelder clyde"),
    ("essential malariology",                                            "bruce chwatt gilles warhurst"),

    # ── Parasitology ───────────────────────────────────────────────────────────
    ("foundations of parasitology",                                      "roberts janovy nadler"),
    ("medical parasitology",                                             "markell"),
    ("human parasitology",                                               "bogitsh"),
    ("parasitic diseases",                                               "despommier"),
    ("atlas of human parasitology",                                      "ash orihel"),
    ("medical helminthology",                                            None),            # various eds.
    ("medical entomology for students",                                  "service"),
    ("vectors of human disease",                                         "busvine"),
    ("mosquitoes of the world",                                          "darsie ward"),
    ("biology of disease vectors",                                       "marquardt"),

    # ── Bacteriology ──────────────────────────────────────────────────────────
    ("the germ theory of disease",                                       "pasteur koch"),
    ("principles of bacteriology",                                       "mackie mccartney topley wilson"),
    ("medical microbiology",                                             "murray"),
    ("bacteriology and immunity",                                        "topley wilson"),
    ("the bacteria",                                                     "gunsalus stanier"),

    # ── Virology ──────────────────────────────────────────────────────────────
    ("medical virology",                                                 "white fenner"),
    ("the pathogenesis of viral infections",                             "galasso merigan buchanan"),
    ("fields virology",                                                  "fields howley knipe"),
    ("an inquiry into the causes and effects of the variolae vaccinae",  "jenner"),

    # ── Immunology ────────────────────────────────────────────────────────────
    ("cellular and molecular immunology",                                "abbas lichtman pillai"),
    ("the immune system",                                                "parham"),

    # ── Narrative / social medicine ───────────────────────────────────────────
    ("and the band played on",                                           "shilts"),

    # ── Internal medicine & clinical ──────────────────────────────────────────
    ("harrisons principles of internal medicine",                        "harrison loscalzo"),
    ("harrison s principles of internal medicine",                       "harrison loscalzo"),
    ("cecil textbook of medicine",                                       "cecil goldman schafer"),
    ("principles and practice of medicine",                              "osler"),
    ("grays anatomy",                                                    "gray standring"),
    ("gray s anatomy",                                                   "gray standring"),
    ("atlas of human anatomy",                                           "netter"),
    ("robbins pathologic basis of disease",                              "robbins kumar"),
    ("oxford handbook of clinical medicine",                             "longmore"),
    ("clinical examination",                                             "macleod"),

    # ── Pathology ─────────────────────────────────────────────────────────────
    ("molecular pathology",                                              "coleman tsongalis"),
    ("diagnostic pathology",                                             "coulibaly"),

    # ── Ethics ────────────────────────────────────────────────────────────────
    ("principles of biomedical ethics",                                  "beauchamp childress"),
    ("ethics and professionalism in medicine",                           None),            # various eds.
]

OUTBREAK_TIMELINE = [
    {"event": "First smallpox vaccination (Jenner)",
     "keywords": ["jenner", "variolae vaccinae", "smallpox vaccination", "cowpox inoculation"],
     "barnard_prefixes": ["KI"]},

    {"event": "Broad Street cholera outbreak / Cholera",
     "keywords": ["cholera", "john snow", "broad street", "vibrio cholerae", "el tor"],
     "barnard_prefixes": ["JK"]},

    {"event": "Discovery of TB bacteria (Koch) / Tuberculosis",
     "keywords": ["tuberculosis", "tubercle bacillus", "robert koch", "koch's bacillus",
                  "mycobacterium tuberculosis"],
     "barnard_prefixes": ["JC"]},

    {"event": "Mosquito transmission of Malaria (Ross) / Malaria",
     "keywords": ["malaria", "ronald ross", "anopheles", "mosquito transmission of malaria",
                  "plasmodium"],
     "barnard_prefixes": ["LF", "NO", "N", "ND", "NC"]},

    {"event": "Yellow fever transmission (Walter Reed) / Yellow fever",
     "keywords": ["yellow fever", "walter reed", "fievre jaune", "fiebre amarilla"],
     "barnard_prefixes": ["KPA"]},

    {"event": "Discovery of Chagas disease",
     "keywords": ["chagas", "trypanosoma cruzi", "doenca de chagas", "enfermedad de chagas"],
     "barnard_prefixes": ["LP", "LN"]},

    {"event": "1918 Influenza pandemic / Influenza",
     "keywords": ["influenza", "spanish flu", "1918 pandemic", "grippe"],
     "barnard_prefixes": ["KL"]},

    {"event": "Malaria eradication campaigns",
     "keywords": ["malaria eradication", "global malaria eradication",
                  "eradication of malaria", "ddt malaria", "malaria campaign"],
     "barnard_prefixes": ["LF", "NO", "N", "ND", "NC"]},

    {"event": "Global smallpox eradication programme / Smallpox",
     "keywords": ["smallpox", "variola", "smallpox eradication", "pocken"],
     "barnard_prefixes": ["KI"]},

    {"event": "HIV/AIDS",
     "keywords": ["hiv", "aids", "sida", "human immunodeficiency virus",
                  "acquired immunodeficiency", "acquired immune deficiency",
                  "virus de l'immunodeficience"],
     "barnard_prefixes": ["KRC", "KR"]},

    {"event": "Global expansion of Dengue",
     "keywords": ["dengue"],
     "barnard_prefixes": ["KPD", "KP"]},

    {"event": "Mpox outbreaks",
     "keywords": ["mpox", "monkeypox"],
     "barnard_prefixes": ["KI"]},

    {"event": "Ebola outbreak in West Africa / Hemorrhagic fever",
     "keywords": ["ebola", "hemorrhagic fever", "haemorrhagic fever", "marburg"],
     "barnard_prefixes": ["KPH"]},

    {"event": "COVID-19 pandemic",
     "keywords": ["covid", "sars-cov", "coronavirus pandemic"],
     "barnard_prefixes": ["K"]},
]

BARNARD = {
    "A":"Reference","AA":"Encyclopedias","AB":"Scientific and medical dictionaries","AC":"Directories","AR":"Education, Teaching",
    "B":"Natural Science","BA":"Scientific literature","BB":"Statistics","BC":"Chemistry","BCB":"Biochemistry","BCC":"Clinical chemistry",
    "BD":"Nutrition","BI":"Microscopy","BJ":"Biology","BJC":"Molecular biology","BJS":"Evolution","BJT":"Genetics","BJX":"Ecology",
    "BK":"Botany","BL":"Zoology","BM":"Anatomy","BN":"Histology","BQ":"Physiology",
    "C":"Internal Medicine","CBA":"Medical education","CBE":"Medical ethics","CBN":"General practice","CBP":"Group practice","CBX":"Nursing","CZ":"Traditional medicine",
    "D":"History of Medicine",
    "E":"Epidemiology","EA":"Disease ecology","EB":"Epidemiology","EC":"Disease surveillance","EH":"Demography","EK":"Mortality","EL":"Mortality","EP":"Medical geography","EV":"Human ecology",
    "G":"Toxicology","GK":"Food poisoning","GL":"Venomous animals",
    "H":"Immunology, Infectious diseases","HD":"Immunology","HG":"Immunodiagnosis","HI":"Vaccination","HW":"Infectious diseases",
    "I":"Mycology",
    "J":"Bacteriology","JC":"Tuberculosis","JD":"Leprosy","JI":"Gonorrhea","JJ":"Bacterial meningitis","JK":"Cholera","JM":"Tetanus","JN":"Plague","JP":"Brucellosis","JS":"Salmonelloses","JX":"Treponematoses","JYC":"Lyme disease","JZ":"Leptospiroses",
    "K":"Virology","KA":"Rickettsioses","KG":"Chlamydia","KGJ":"Trachoma","KI":"Smallpox","KK":"Rabies","KP":"Arboviral diseases","KPA":"Yellow fever","KPD":"Dengue","KPH":"Hemorrhagic fever","KQ":"Encephalitis","KR":"Retroviruses","KRC":"HIV, AIDS","KS":"Hepatitis","KT":"Poliomyelitis",
    "L":"Parasitology","LA":"Protozoology","LC":"Babesiasis","LD":"Coccidiasis","LDC":"Cryptosporidiasis","LDP":"Pneumocystosis","LE":"Toxoplasmosis","LF":"Malaria","LL":"Amebiasis","LN":"Sleeping sickness","LNX":"Animal trypanosomiasis","LP":"Chagas disease","LQ":"Leishmaniasis","LV":"Giardiasis",
    "M":"Helminthology","MH":"Schistosomiasis","MK":"Echinococcosis","MQ":"Ascariasis","MQS":"Strongyloidiasis","MR":"Ankylostomiasis","MS":"Filariases","MU":"Onchocerciasis","MV":"Trichinelliasis",
    "N":"Entomology","NB":"Ectoparasites","NC":"Vector-borne diseases","ND":"Vector control","NDI":"Insecticides","NDY":"Biological control","NG":"Mites","NGT":"Ticks","NK":"Bugs","NL":"Flies","NO":"Mosquitoes","NP":"Simuliidae","NT":"Tsetse flies",
    "O":"Intermediate Hosts",
    "P":"Pathology","PM":"Cancer","PY":"Hematology","PYB":"Blood groups","PYE":"Blood transfusion","PYH":"Anemia","PYHR":"Hemoglobinopathies","PYHS":"Sickle cell anemia","PYK":"Nutritional anemias",
    "Q":"Diagnosis","QD":"Clinical medicine","QH":"Laboratory medicine","QM":"Laboratory medicine","QR":"Radiology",
    "R":"Pharmacology","RJ":"Pharmacology","RN":"Therapeutics","RNA":"Clinical trials","RNR":"Drug resistance","RS":"Drug supply",
    "S":"Public Health","SA":"Environmental hygiene","SD":"Water hygiene","SJA":"Disaster relief","SO":"Public health","SOA":"Health organizations","SOB":"Health information","SOC":"Health evaluation","SOCA":"Health research","SOD":"Health economics","SOE":"Health financing","SOF":"Health manpower","SOG":"Hospitals","SOH":"Health education","SON":"Medical sociology","SOP":"Migrants health","SP":"Preventive medicine","SPA":"Disinfection","SPZ":"Hospital infections","SS":"Travel health","ST":"Urban health","SW":"Naval health",
    "U":"Medical Specialties","UB":"Tropical medicine","UBR":"Tropical medicine","UH":"Cardiology","UI":"Neurology","UJ":"Psychiatry","UK":"Ophthalmology","UL":"Oto-rhino-laryngology","UO":"Respiratory diseases","UP":"Gastroenterology","UPT":"Enteric infections","UQ":"Endocrinology","UR":"Dermatology","US":"Urology","UT":"Sexology","UTH":"Fertility","UTT":"Family planning","UTU":"Sexually transmitted diseases","UV":"Gynecology","UW":"Obstetrics","UWB":"Midwives","UWE":"Mother and child health","UX":"Pediatrics","UY":"Geriatrics",
    "V":"Surgery","VG":"Anesthesia","W":"Oral Health",
    "X":"Veterinary Sciences","XC":"Veterinary medicine","XE":"Epizootiology","XH":"Immunology","XHW":"Infectious diseases","XJ":"Microbiology","XK":"Virology","XL":"Parasitology","XM":"Helminthology","XN":"Entomology","XO":"Zoonoses","XS":"Hygiene","XW":"Animal husbandry","XXB":"Cattle","XXN":"Small ruminants","XXO":"Pigs","XXU":"Poultry","XXZ":"Wild animals","XZ":"Laboratory animals",
    "Y":"Agriculture","YA":"Agriculture","YC":"Field crops","YH":"Pest control",
    "Z":"Geography, Sociology","ZB":"Geography","ZVM":"Development","ZVO":"Economics","ZVP":"Political Science","ZVU":"Management","ZVY":"Colonial administration","ZY":"Population issues","ZYG":"Population movement",
}

RETENTION_FLAGS = {
    "A":"H3","AA":"H3","AB":"H3","AC":"H3","AR":"H2",
    "BA":"H2","BB":"H2","BC":"H3","BCB":"H3","BCC":"H3","BD":"H2","BE":"H3","BF":"H3",
    "BJ":"H2","BJC":"H2","BJS":"H2","BJT":"H2","BJX":"H2","BK":"H2","BL":"H2","BM":"H3","BN":"H2","BQ":"H2",
    "C":"H2","CBA":"H2","CBE":"H1","CBN":"H3","CBP":"H3","CBX":"H3","CZ":"H1",
    "D":"H1",
    "EA":"H2","EB":"H1","EC":"H2","EH":"H2","EK":"H2","EL":"H2","EP":"H2","EQ":"H2","ES":"H2","ET":"H2","EV":"H2",
    "G":"H2","GK":"H3","GL":"H2",
    "H":"H2","HD":"H2","HG":"H2","HI":"H2","HW":"H1",
    "I":"H2",
    "J":"H2","JC":"H1","JD":"H2","JI":"H2","JJ":"H2","JK":"H2","JM":"H2","JN":"H2","JP":"H2","JS":"H2","JX":"H2","JYC":"H2","JZ":"H2",
    "K":"H2","KA":"H2","KG":"H2","KGJ":"H2","KI":"H2","KK":"H2","KP":"H2","KPA":"H2","KPD":"H2","KPH":"H2","KQ":"H2","KR":"H2","KRC":"H2","KS":"H2","KT":"H2",
    "L":"H2","LA":"H2","LC":"H2","LD":"H2","LDC":"H2","LDP":"H2","LE":"H2","LF":"H1","LL":"H2","LN":"H2","LNX":"H2","LP":"H2","LQ":"H2","LV":"H2",
    "M":"H2","MH":"H2","MK":"H2","MQ":"H2","MQS":"H2","MR":"H2","MS":"H2","MU":"H2","MV":"H2","MW":"H2",
    "N":"H2","NB":"H2","NC":"H2","ND":"H2","NDI":"H3","NDY":"H2","NG":"H2","NGT":"H2","NK":"H2","NL":"H2","NO":"H2","NP":"H2","NT":"H2",
    "O":"H2",
    "P":"H2","PM":"H2","PY":"H2","PYB":"H2","PYE":"H2","PYH":"H2","PYHR":"H2","PYHS":"H2","PYK":"H2",
    "Q":"H3","QA":"H3","QD":"H3","QH":"H3","QM":"H3","QR":"H3",
    "R":"H3","RJ":"H3","RN":"H3","RK":"H3","RNA":"H2","RNR":"H3","RS":"H3",
    "S":"H2","SA":"H2","SB":"H2","SC":"H2","SD":"H2","SJA":"H2","SO":"H2","SOA":"H2","SOB":"H2","SOC":"H2","SOCA":"H2","SOD":"H2","SOE":"H2","SOF":"H2","SOG":"H2","SOH":"H2","SON":"H2","SOP":"H2","SP":"H2","SPA":"H2","SPZ":"H2","SS":"H2","ST":"H2","SW":"H2",
    "U":"H2","UB":"H1","UBR":"H1","UH":"H2","UI":"H2","UJ":"H2","UK":"H2","UL":"H2","UO":"H2","UP":"H2","UPT":"H2","UQ":"H2","UR":"H2","US":"H2","UT":"H2","UTH":"H2","UTT":"H2","UTU":"H2","UV":"H2","UW":"H2","UWB":"H2","UWE":"H2","UX":"H2","UY":"H2",
    "V":"H3","V":"H3","VA":"H3","VB":"H3","VC":"H3","VD":"H3","VE":"H3","VF":"H3","VG":"H3","W":"H3",
    "X":"H2","XC":"H2","XE":"H2","XH":"H2","XHW":"H2","XJ":"H2","XK":"H2","XL":"H2","XM":"H2","XN":"H2","XO":"H2","XS":"H2","XW":"H3","XXB":"H3","XXN":"H3","XXO":"H3","XXU":"H3","XXZ":"H2","XZ":"H2",
    "Y":"H3","YA":"H3","YB":"H3","YC":"H3","YH":"H3",
    "Z":"H2","ZB":"H2","ZI":"H3","ZR":"H3","ZV":"H2","ZVL":"H2","ZVM":"H2","ZVN":"H2","ZVO":"H2","ZVP":"H2","ZVU":"H2","ZVY":"H2","ZY":"H2","ZYG":"H2",
}

CONGO_TERMS   = ["congo","belgian congo","zaire","belgique","belgium","kinshasa","leopoldville","brazzaville","katanga"]
IRCB_TERMS    = ["ircb","institut royal colonial belge","royal belgian colonial institute",
                 "koninklijk belgisch koloniaal instituut","arsc","academie royale des sciences coloniales",
                 "royal academy of colonial sciences","koninklijke academie voor koloniale wetenschappen"]


# ── Helpers ──────────────────────────────────────────────────
def gf(rec, *tags):
    for t in tags:
        v = rec.get(t)
        if v:
            return v[0] if isinstance(v, list) else v
    return ""

def barnard_label(code):
    if not code:
        return ""
    c = code.strip().upper()
    if c in BARNARD:
        return f"{c} — {BARNARD[c]}"
    for i in range(len(c)-1, 0, -1):
        p = c[:i]
        if p in BARNARD:
            return f"{c} — {BARNARD[p]}"
    return c

def base_barnard(code):
    """Normalise Barnard code to base class (first 2-3 chars)."""
    if not code:
        return ""
    c = code.strip().upper()
    return c[:3] if len(c) >= 3 else c

def get_retention_flag(code):
    if not code:
        return None
    c = code.strip().upper()
    if c in RETENTION_FLAGS:
        return RETENTION_FLAGS[c]
    for i in range(len(c)-1, 0, -1):
        p = c[:i]
        if p in RETENTION_FLAGS:
            return RETENTION_FLAGS[p]
    return None

def get_year(rec):
    raw = gf(rec, "Y1", "PY", "DA")
    m = re.search(r"\d{4}", raw)
    return int(m.group()) if m else None

def get_isbn(rec):
    return re.sub(r"[^0-9X]", "", gf(rec, "SN"), flags=re.IGNORECASE)

def get_authors(rec):
    all_authors = []
    for tag in ("A1", "A2"):
        v = rec.get(tag, [])
        all_authors += v if isinstance(v, list) else [v]
    if not all_authors:
        return ""
    if len(all_authors) <= 2:
        return "; ".join(all_authors)
    return "; ".join(all_authors[:2]) + " et al."

def make_circ_key(barnard, call_num, year):
    b = (barnard or "").strip().upper()
    c = re.sub(r"/.*$", "", (call_num or "").strip())
    c = re.sub(r"M$", "", c, flags=re.IGNORECASE).strip()
    y = (year or "").strip()
    return f"{b}|{c}|{y}"

def get_circ_key(rec):
    b = gf(rec, "U4").strip().upper()
    c = re.sub(r"M$", "", gf(rec, "U5").strip(), flags=re.IGNORECASE).strip()
    raw = gf(rec, "Y1", "PY", "DA")
    m = re.search(r"\d{4}", raw)
    y = m.group() if m else ""
    return make_circ_key(b, c, y)

def unicat_url(rec):
    import urllib.parse
    isbn = get_isbn(rec)
    if isbn:
        return f"https://www.unicat.be/uniCat?func=search&query={urllib.parse.quote(isbn)}"
    # Clean title: strip punctuation, brackets, take first 4 meaningful words
    raw_title = gf(rec, "T1", "TI")
    clean_title = re.sub(r"[\[\];:,!?(){}]", " ", raw_title)
    clean_title = re.sub(r"\s+", " ", clean_title).strip()
    title  = " ".join(clean_title.split()[:4])
    author = " ".join(gf(rec, "A1", "A2").split()[:2])
    q = " ".join(filter(None, [title, author]))
    return f"https://www.unicat.be/uniCat?func=search&query={urllib.parse.quote(q)}"

def is_historical_title(title, author=""):
    """Return True if title+author match a known landmark work.

    Each HISTORICAL_TITLES entry is (title_fragment, author_fragment_or_None).
    - title_fragment must appear in the cleaned title AND cover ≥60% of its words.
    - If author_fragment is set, it must also appear in the cleaned author string.
    """
    t = re.sub(r"[^a-z0-9 ]", " ", title.lower())
    t = re.sub(r"\s+", " ", t).strip()
    a = re.sub(r"[^a-z ]", " ", author.lower())
    a = re.sub(r"\s+", " ", a).strip()

    for title_frag, author_frag in HISTORICAL_TITLES:
        if title_frag not in t:
            # Also allow: historical fragment starts with first 20 chars of record title
            if not (len(t) >= 20 and title_frag.startswith(t[:20])):
                continue
        # Coverage guard: fragment must cover ≥60% of the record title's words
        ht_words = len(title_frag.split())
        t_words  = len(t.split())
        if t_words > 0 and (ht_words / t_words) < 0.6:
            continue
        # Author guard — author_frag may contain space-separated alternatives (any one suffices)
        if author_frag is not None:
            alternatives = author_frag.split()
            if not any(alt in a for alt in alternatives):
                continue
        return True
    return False

def matches_outbreak(title, abstract, keywords, barnard):
    # Keywords field excluded from haystack — auto-assigned from abstract and unreliable.
    # Matching on title + abstract only.
    # Each event also requires the record's Barnard class to start with one of its
    # barnard_prefixes — this prevents cross-domain false positives (e.g. a pediatrics
    # textbook matching "tuberculosis" because of an incidental abstract mention).
    raw_hay = (title + " " + abstract).lower()
    # Strip accents so "fièvre" matches keyword "fievre", "doença" matches "doenca", etc.
    hay = unicodedata.normalize("NFKD", raw_hay).encode("ascii", "ignore").decode("ascii")
    barnard_upper = (barnard or "").upper()
    matched = []
    for ev in OUTBREAK_TIMELINE:
        if not any(k in hay for k in ev["keywords"]):
            continue
        prefixes = ev.get("barnard_prefixes", [])
        if prefixes and not any(barnard_upper.startswith(p) for p in prefixes):
            continue
        matched.append(ev["event"])
    return matched


# ── RIS parser ───────────────────────────────────────────────
def parse_ris(text):
    records = []
    for block in re.split(r"\nER\s*-", text):
        rec = {}
        for line in block.strip().splitlines():
            m = re.match(r"^([A-Z][A-Z0-9])\s*-\s*(.*)$", line)
            if not m:
                continue
            tag, val = m.group(1), m.group(2).strip()
            if tag in rec:
                if isinstance(rec[tag], list):
                    rec[tag].append(val)
                else:
                    rec[tag] = [rec[tag], val]
            else:
                rec[tag] = val
        if rec.get("T1") or rec.get("TI"):
            records.append(rec)
    return records


# ── Circulation loader ───────────────────────────────────────
def load_circulation(path, delimiter, b_col=1, c_col=2, y_col=3):
    borrowed = set()
    try:
        with open(path, encoding="utf-8-sig", errors="replace") as f:
            for line in f:
                cols = line.rstrip("\n").split(delimiter)
                if len(cols) <= max(b_col, c_col, y_col):
                    continue
                barnard  = cols[b_col].strip()
                call_num = cols[c_col].strip()
                year     = cols[y_col].strip()
                if barnard and call_num and re.search(r"\d", call_num):
                    borrowed.add(make_circ_key(barnard, call_num, year))
    except FileNotFoundError:
        print(f"  Warning: file not found: {path}", file=sys.stderr)
    return borrowed


# ── Rule engine ──────────────────────────────────────────────
def apply_rules(rec, all_records, borrowed_bibs, isbn_counts, older_edition=False, translation_duplicate=False, barnard_counts=None):
    year          = get_year(rec)
    isbn          = get_isbn(rec)
    rec_type      = gf(rec, "TY").upper()
    title         = gf(rec, "T1", "TI")
    abstract      = gf(rec, "N2")
    kw_list       = rec.get("KW", [])
    keywords      = ", ".join(kw_list if isinstance(kw_list, list) else [kw_list])
    publisher     = gf(rec, "PB").lower()
    note          = gf(rec, "N1")
    doc_type      = gf(rec, "U1")
    e_status      = gf(rec, "U2")
    barnard       = gf(rec, "U4")
    retention     = get_retention_flag(barnard)
    has_e_version = ("e-also" in e_status.lower() or "-h" in doc_type.lower() or bool(gf(rec, "L2")))

    flags         = []
    keep_override = False
    hard_rule     = None
    had_keep_reason = False  # tracks if any keep rule fired, even if later overridden

    def flag(criterion, detail, severity):
        flags.append({"criterion": criterion, "detail": detail, "severity": severity})

    # ── E-only skip ──
    if "e-only" in e_status.lower() or "-e" in doc_type.lower():
        flag("E-only", "No physical copy", "keep")
        return {"recommendation": "SKIP", "keep_override": False, "hard_rule": "SKIP",
                "flags": flags, "retention": retention, "historically_significant": False,
                "historical_reasons": [], "reasoning": "Skipped — e-only."}

    # ── H1 class ──
    if retention == "H1":
        flag("Retention flag", f"H1 — Core historical field ({barnard_label(barnard)}) — retain indefinitely", "keep")
        keep_override = True
        had_keep_reason = True

    # ── Dissertation ──
    if (rec_type in ("THES","THESIS","DISS")
            or "dissertation" in doc_type.lower()
            or "dissertation" in title.lower()):
        flag("Document type", "Doctoral dissertation — always keep", "keep")
        keep_override = True
        had_keep_reason = True

    # ── Proceedings ──
    CONFERENCE_TERMS = [
        "proceedings", "conference", "congres", "congrès", "congress", "congresso",
        "symposium", "colloque", "workshop", "abstracts", "anais", "arquivos",
        "conférence", "conferencia", "conferência", "meeting", "congresos",
        "congresso", "colloqui", "seminaire", "séminaire",
        "kongress", "konferenz", "tagung", "jahrestagung", "symposia",
    ]
    TROPICAL_TERMS = [
        "tropical", "malaria", "leprosy", "leprol", "lèpre", "lepra", "lepre", "parasit",
        "infectious", "infecti", "colonial", "africa", "afrique", "african",
        "congo", "asia", "latin america", "developing", "health", "santé", "sante",
        "hygiene", "hygiène", "epidemiol", "immunol", "bacteriol", "virol",
        "entomol", "helminth", "schistosom", "trypanosoma", "filaria",
        "microbiolog", "chemotherap", "pharmacol", "nutrition", "medic",
        "tropenmedizin", "tropen", "geneeskunde", "gezondheid", "salud",
        "coccidi", "protozoa", "helminth", "nematod", "cestod", "trematod",
    ]
    is_proceedings = (
        rec_type in ("CONF","CPAPER")
        or any(t in title.lower() for t in CONFERENCE_TERMS)
    )
    if is_proceedings:
        hay_conf = (title + " " + abstract + " " + keywords).lower()
        is_tropical_conf = any(t in hay_conf for t in TROPICAL_TERMS)
        if not has_e_version:
            if is_tropical_conf:
                flag("Document type", "Conference proceedings on tropical/colonial topic — always keep", "keep")
                keep_override = True
                had_keep_reason = True
            else:
                flag("Document type", "Conference proceedings (no e-version) — keep", "keep")
                keep_override = True
                had_keep_reason = True
        else:
            flag("Document type", "Conference proceedings — e-version exists", "weed")
            hard_rule = "WEED"

    # ── Liber amicorum ──
    # Liber amicorum, festschrift, dedications, commemorative editions
    dedication_terms = [
        "liber amicorum", "festschrift",
        # Dedications to a person
        "dédié à", "dedie a", "dedicated to", "ter ere van", "ter gelegenheid van",
        "à l'occasion de", "a l'occasion de", "bij gelegenheid",
        "hommage à", "hommage a", "in honour of", "in honor of",
        "in memoriam", "à la mémoire de", "à la memoire de",
        # Commemorative / anniversary editions
        "anniversaire", "anniversary", "jubileum", "jubilé", "jubile",
        "commemorat", "commémorat", "gedenkboek", "gedenkschrift",
        "centenaire", "centenario", "centenário", "centennial", "centenary",
        "comemorativ", "conmemorat",
    ]
    hay_ded = (title + " " + abstract).lower()
    if any(t in hay_ded for t in dedication_terms):
        flag("Document type", "Liber amicorum / dedication / commemorative edition — always keep", "keep")
        keep_override = True
        had_keep_reason = True

    # ── Congo / Belgium ──
    hay = (title + " " + abstract + " " + publisher + " " + keywords).lower()
    if any(t in hay for t in CONGO_TERMS):
        flag("Regional relevance", "Congo / Belgium — always keep", "keep")
        keep_override = True
        had_keep_reason = True

    # ── Africa-specific studies — local/country-level research ──
    AFRICA_TERMS = [
        "africa", "african", "afrique", "africain",
        # specific countries
        "ghana", "gambia", "gambian", "nigeria", "nigerian", "kenya", "kenyan",
        "tanzania", "tanzanian", "uganda", "ugandan", "ethiopia", "ethiopian",
        "senegal", "senegalese", "mali", "malian", "niger", "cameroon", "cameroonian",
        "mozambique", "zimbabwe", "zambia", "malawi", "rwanda", "burundi",
        "somalia", "sudan", "south africa", "namibia", "botswana", "lesotho",
        "sierra leone", "liberia", "guinea", "ivory coast", "côte d'ivoire",
        "cote d ivoire", "togo", "benin", "burkina", "angola", "madagascar",
        "mauritius", "reunion", "egypt", "nigeria", "chad", "tchad",
        "abidjan", "dakar", "nairobi", "kampala", "dar es salaam", "lagos",
        "accra", "bamako", "ouagadougou", "niamey", "lome", "cotonou",
    ]
    # Also keep atlases regardless of region — geographic atlases are always scarce
    is_atlas = "atlas" in title.lower() or "atlas" in (gf(rec, "U5") or "").lower()
    hay_africa = (title + " " + abstract + " " + keywords + " " + gf(rec, "CY")).lower()
    is_africa_specific = any(t in hay_africa for t in AFRICA_TERMS)

    if is_atlas and is_africa_specific:
        flag("Regional relevance", "African atlas — geographically scarce, always keep", "keep")
        keep_override = True
        had_keep_reason = True
    elif is_africa_specific and not any(t in hay_africa for t in ["manual", "guide", "handbook", "directory"]):
        # Africa-specific non-manual studies → keep for historical/regional value
        flag("Regional relevance", "Africa-specific study — historical/regional value, keep", "keep")
        keep_override = True
        had_keep_reason = True

    # ── IRCB / ARSC ──
    hay2 = (title + " " + abstract + " " + publisher + " " + keywords + " " + note).lower()
    if any(t in hay2 for t in IRCB_TERMS):
        flag("IRCB/ARSC", "IRCB or ARSC publication — always keep", "keep")
        keep_override = True
        had_keep_reason = True

    # ── Historical title list ──
    author_str = get_authors(rec)
    if is_historical_title(title, author_str):
        flag("Historical title", "Listed in Top Historical Titles for Tropical Medicine", "keep")
        keep_override = True
        had_keep_reason = True

    # ── Outbreak relevance ──
    outbreaks = matches_outbreak(title, abstract, keywords, barnard)
    if outbreaks:
        flag("Outbreak relevance", "Relates to: " + "; ".join(outbreaks[:3]), "keep")
        keep_override = True
        had_keep_reason = True

    # ── Circulation ──
    if borrowed_bibs and get_circ_key(rec) in borrowed_bibs:
        flag("Circulation", "Borrowed at least once (2019–2026) — keep", "keep")
        keep_override = True
        had_keep_reason = True

    # ── Archives-G location — curated special collection ──
    u2 = gf(rec, 'U2').lower()
    if 'archives-g' in u2:
        flag("Archives-G", "Stored in Archives-G — curated special collection, always keep", "keep")
        keep_override = True
        had_keep_reason = True

    # ── Specialist tropical medicine publisher ──
    SPECIALIST_PUBLISHERS = [
        # ITG / ITM (Antwerp)
        "instituut voor tropische geneeskunde", "institut de medecine tropicale",
        "institute of tropical medicine", "prins leopold instituut", "prince leopold institute",
        "institut de medecine tropicale prince",
        # Other tropical medicine schools
        "london school of hygiene and tropical medicine", "london school of hygiene",
        "liverpool school of tropical medicine",
        "bernhard-nocht-institut", "swiss tropical institute",
        "mahidol university, faculty of tropical medicine",
        # African research institutes
        "musee royal de l", "musee royal du congo", "royal museum for central africa",
        "academie royale des sciences coloniales", "academie royale des sciences d",
        "institut royal colonial belge", "arsom", "arsc", "ircb",
        "orstom", "iemvt", "occge", "oceac", "fometro", "berps",
        "international livestock centre for africa", "ilca",
        # International tropical disease programmes
        "undp/world bank/who special programme", "unicef/undp/world bank/who",
        "special programme for research and training in tropical diseases",
        "division of control of tropical diseases",
        # Veterinary tropical medicine
        "centre for tropical veterinary medicine",
        "medecine veterinaire des pays tropicaux",
        "tropical products institute", "tropical health technology",
        # Other specialist
        "royal society of tropical medicine",
        "american society of tropical medicine", "bureau of hygiene and tropical diseases",
        "ross institute of tropical hygiene",
        # KIT Amsterdam (Royal Tropical Institute)
        "koninklijk instituut voor de tropen", "royal tropical institute",
        "institut royal des tropiques", "kit amsterdam",
    ]
    pub_lower = gf(rec, 'PB').lower()
    a3_vals = rec.get('A3', [])
    a3_vals = [a3_vals] if isinstance(a3_vals, str) else a3_vals
    a3_lower = ' '.join(a3_vals).lower()
    hay_pub = pub_lower + ' ' + a3_lower

    if not keep_override and any(sp in hay_pub for sp in SPECIALIST_PUBLISHERS):
        flag("Specialist publisher", f"Published by specialist tropical medicine institution — keep", "keep")
        keep_override = True
        had_keep_reason = True

    # ── Barnard (informational) ──
    if barnard:
        flag("Barnard", barnard_label(barnard) + (f" [{retention}]" if retention else ""), "keep")

    # ── WHO/FAO open access downgrade — KEEP → REVIEW ──
    WHO_FAO_TERMS = ["world health organization", "world health organisation",
                     "food and agriculture organization", "food and agriculture organisation",
                     "wereldgezondheidsorganisatie", "organisation mondiale de la sante",
                     "organisation mondiale de la santé", "organización mundial de la salud"]
    WHO_FAO_SHORT = ["who", "fao", "oms", "oas"]  # short codes — match as whole words only
    # Check all possible author/publisher/editor fields
    # Collect all author/publisher fields including list types (A3 is often corporate author)
    _pub_parts = [publisher]
    for tag in ("A1", "A2", "A3", "A4", "ED", "SE", "PB", "T3"):
        v = rec.get(tag, "")
        if isinstance(v, list):
            _pub_parts.extend(v)
        elif v:
            _pub_parts.append(v)
    hay_pub = " ".join(filter(None, _pub_parts)).lower()
    is_who_fao = (
        any(t in hay_pub for t in WHO_FAO_TERMS)
        or any(re.search(r"\b" + t + r"\b", hay_pub) for t in WHO_FAO_SHORT)
    )
    if is_who_fao:
        flag("WHO/FAO", "WHO or FAO publication — likely available open access online — verify before keeping", "review")

    # ── Translation duplicate — non-English version when English exists ──
    # The English copy is the one to retain. If a non-English copy of the same work
    # also exists in English elsewhere in the collection, weed the non-English copy
    # outright — this overrides any keep rule that may have fired on it (H1 class,
    # outbreak relevance, etc.), since the English edition already covers that
    # content and is the preferred copy to keep. The English edition itself never
    # carries the translation_duplicate flag, so it is unaffected either way.
    if translation_duplicate:
        keep_override = False
        flag("Translation", "Non-English version — English edition exists in collection — weed "
                             "(English copy is retained as the preferred edition)", "weed")
        hard_rule = "WEED"

    # ── Older edition rule — superseded by newer edition in same collection ──
    if older_edition and not keep_override:
        flag("Edition", "Older edition — newer edition exists in collection — weed", "weed")
        hard_rule = "WEED"

    # ── Barnard class scarcity protection ──
    # If a Barnard class has fewer than 3 items in the whole collection,
    # don't weed unless it's an outdated manual or clinical guideline
    if not keep_override and barnard_counts is not None:
        bc = base_barnard(barnard) if barnard else ""
        class_count = barnard_counts.get(bc, 0) if bc else 0
        MANUAL_GUIDE_TERMS = ["manual", "guide", "handbook", "guideline", "directory",
                               "formulary", "protocol", "procedure", "curriculum"]
        is_manual_guide = any(t in title.lower() for t in MANUAL_GUIDE_TERMS)
        if class_count <= 2 and not is_manual_guide:
            flag("Class scarcity", f"Only {class_count} item(s) in Barnard class {bc} — keep to avoid collection gap", "keep")
            keep_override = True
            had_keep_reason = True

    # ── Weed rules (only if not already kept) ──
    if not keep_override:
        # Duplicate
        if isbn and isbn_counts.get(isbn, 0) > 1:
            flag("Duplicate", f"{isbn_counts[isbn]} copies — this is a duplicate", "weed")
            hard_rule = "WEED"

        if year is not None:
            if 1950 <= year <= 1990:
                if retention == "H3":
                    flag("Publication date", f"Published {year} [H3 class] — auto-weed", "weed")
                    hard_rule = "WEED"
                elif retention == "H2":
                    flag("Publication date", f"Published {year} [H2 class] — auto-weed (no special flags)", "weed")
                    hard_rule = "WEED"
                else:
                    flag("Publication date", f"Published {year} (1950–1990)", "weed")
                    hard_rule = "WEED"
            elif year < 1950:
                flag("Publication date", f"Published {year} (pre-1950) — potential historical value", "keep")
                keep_override = True
                had_keep_reason = True
            elif year > 1990:
                flag("Publication date", f"Published {year} (post-1990, not circulated) — auto-weed", "weed")
                hard_rule = "WEED"
        else:
            flag("Publication date", "No publication year — flagged for manual review", "review")
            hard_rule = "REVIEW"

    # ── WHO/FAO final check — must run last so nothing overwrites REVIEW ──
    # Fires if WHO/FAO AND a keep rule fired at some point (even if later overridden by date rule).
    # Skipped for translation duplicates — those are already a final WEED decision
    # (the English edition is the retained copy and already carries any WHO/FAO note).
    if is_who_fao and (keep_override or had_keep_reason) and not translation_duplicate:
        keep_override = False
        hard_rule = "REVIEW"

    recommendation = "KEEP" if keep_override else (hard_rule or "REVIEW")
    hist_criteria  = {"Historical title","Outbreak relevance","Regional relevance","IRCB/ARSC","Retention flag"}
    historically   = (keep_override or recommendation == "KEEP") and any(f["criterion"] in hist_criteria for f in flags)
    hist_reasons   = [f["detail"] for f in flags if f["severity"] == "keep"]
    reasoning      = ". ".join(f["detail"] for f in flags)

    return {
        "recommendation":        recommendation,
        "keep_override":         keep_override,
        "hard_rule":             hard_rule,
        "flags":                 flags,
        "retention":             retention,
        "historically_significant": historically,
        "historical_reasons":    hist_reasons,
        "reasoning":             reasoning,
    }


# ── XLSX export ──────────────────────────────────────────────
def export_xlsx(rows, out_path):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl not found — installing...", file=sys.stderr)
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

    headers = [
        "Title","Author","Year","Type","Bib#","ISBN",
        "Barnard","Retention Flag","Call Number","Language","Location",
        "Recommendation","Circulated","Historically Significant",
        "Historical Reasons","Triggered Rules","Check UniCat","Reasoning",
    ]
    col_widths = [50,25,6,8,10,14,30,8,12,6,14,14,10,12,40,60,50,60]

    fill_map   = {"WEED": "FFDDDD", "KEEP": "DDFFDD", "REVIEW": "FFFFCC", "SKIP": "EEEEEE"}
    header_fill       = PatternFill("solid", fgColor="1E3A5F")
    header_fill_dept  = PatternFill("solid", fgColor="4A235A")  # purple for department sheet

    def make_row_values(row):
        result = row["result"]
        flags  = result["flags"]
        rec    = row["rec"]
        circulated = any(f["criterion"] == "Circulation" for f in flags)
        triggered  = "; ".join(f"{f['criterion']}: {f['detail']}" for f in flags)
        rec_val    = result["recommendation"]
        return rec_val, [
            row["title"], row["author"], row["year"], row["rec_type"],
            gf(rec, "ID"), row["isbn"],
            barnard_label(gf(rec, "U4")), result["retention"] or "",
            gf(rec, "U5"), gf(rec, "U3"), row["location"],
            rec_val,
            "Yes" if circulated else "No",
            "Yes" if result["historically_significant"] else "No",
            "; ".join(result["historical_reasons"]),
            triggered,
            "YES" if (result["recommendation"] == "WEED" and row["isbn"]) else "",
            result["reasoning"],
        ]

    def write_sheet(ws, sheet_rows, hdr_fill):
        # Header
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = hdr_fill
            cell.alignment = Alignment(wrap_text=True)
        # Data
        for excel_row, row in enumerate(sheet_rows, 2):
            rec_val, values = make_row_values(row)
            for ci, v in enumerate(values, 1):
                cell = ws.cell(row=excel_row, column=ci, value=v)
                cell.fill = PatternFill("solid", fgColor=fill_map.get(rec_val, "FFFFFF"))
                cell.alignment = Alignment(wrap_text=True)
        # Widths / freeze / filter
        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.freeze_panes = "A2"
        if ws.max_row > 1:
            ws.auto_filter.ref = ws.dimensions

    # Split rows into library vs department
    lib_rows  = [r for r in rows if "dep-a" not in r["location"].lower()]
    dept_rows = [r for r in rows if "dep-a" in r["location"].lower()]

    wb = Workbook()
    ws_lib = wb.active
    ws_lib.title = "Library Collection"
    write_sheet(ws_lib, lib_rows, header_fill)

    ws_dept = wb.create_sheet("Department Books")
    write_sheet(ws_dept, dept_rows, header_fill_dept)

    print(f"  Sheet 'Library Collection': {len(lib_rows):,} records")
    print(f"  Sheet 'Department Books':   {len(dept_rows):,} records")

    wb.save(out_path)


# ── Main ─────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Library Weeding Agent — ITG Antwerp")
    parser.add_argument("ris",               help="Path to .ris file(s) — space-separated for multiple files", nargs="+")
    parser.add_argument("--students",        help="Student loans CSV (semicolon-delimited)")
    parser.add_argument("--staff",           help="Staff loans CSV/TSV (tab-delimited)")
    parser.add_argument("--out",             default="data/output/weeding_report9.xlsx", help="Output XLSX filename")
    args = parser.parse_args()

    # Load RIS
    def read_ris_file(path):
        for enc in ("utf-8-sig", "utf-8", "cp850", "windows-1252"):
            try:
                text = Path(path).read_text(encoding=enc)
                if "�" not in text:
                    print(f"  Encoding detected: {enc}")
                    return text
            except (UnicodeDecodeError, LookupError):
                continue
        text = Path(path).read_text(encoding="cp850", errors="replace")
        print("  Encoding detected: cp850 (fallback)")
        return text

    all_records = []
    primary_count = 0  # records from first file only (to be processed)
    for file_idx, ris_path in enumerate(args.ris):
        print(f"Reading {ris_path}...")
        ris_text = read_ris_file(ris_path)
        recs = parse_ris(ris_text)
        print(f"  Parsed {len(recs):,} records")
        all_records.extend(recs)
        if file_idx == 0:
            primary_count = len(recs)
    records = all_records
    if len(args.ris) > 1:
        print(f"  Total: {len(records):,} records across {len(args.ris)} file(s)")
        print(f"  Processing first file only ({primary_count:,} records); rest used for collection stats")

    # Load circulation data
    borrowed = set()
    if args.students:
        print(f"Loading student loans: {args.students}")
        s = load_circulation(args.students, delimiter=";")
        borrowed |= s
        print(f"  {len(s):,} student loan records")
    if args.staff:
        print(f"Loading staff loans: {args.staff}")
        s = load_circulation(args.staff, delimiter="\t")
        borrowed |= s
        print(f"  {len(s):,} staff loan records")
    if borrowed:
        print(f"  {len(borrowed):,} unique circulated items total")

    # Pre-compute ISBN counts for duplicate detection
    isbn_counts = {}
    for rec in records:
        isbn = get_isbn(rec)
        if isbn:
            isbn_counts[isbn] = isbn_counts.get(isbn, 0) + 1

    # Pre-compute Barnard class counts — for scarcity protection
    barnard_counts = {}
    for rec in records:
        bc = base_barnard(gf(rec, "U4"))
        if bc:
            barnard_counts[bc] = barnard_counts.get(bc, 0) + 1

    # Pre-compute edition groups: normalise title, group by base title + author
    # For each group, find the record with the highest edition number / latest year
    # so older editions can be flagged as superseded
    import re as _re

    def normalise_title(title):
        """Strip edition info and punctuation for grouping."""
        t = title.lower()
        # Remove edition markers: "3rd ed", "4th edition", "2e éd", etc.
        t = _re.sub(r"[;,]\s*(\d+(st|nd|rd|th)?\.?\s*(ed|edition|éd|uitgave|druk|aufl).*)", "", t)
        t = _re.sub(r"\s*(\d+(st|nd|rd|th)?\.?\s*(ed|edition|éd|uitgave|druk|aufl).*)", "", t)
        t = _re.sub(r"[^a-z0-9 ]", " ", t)
        t = _re.sub(r"\s+", " ", t).strip()
        return t[:60]  # first 60 chars as key

    def get_edition_num(rec):
        """Extract edition number from title, defaulting to 1."""
        title = gf(rec, "T1", "TI")
        m = _re.search(r"(\d+)(st|nd|rd|th)?\s*(ed|edition|éd|uitgave|druk|aufl)", title, _re.I)
        if m:
            return int(m.group(1))
        return 1

    # Group records by normalised title
    from collections import defaultdict
    edition_groups = defaultdict(list)
    for i, rec in enumerate(records):
        title = gf(rec, "T1", "TI")
        if not title:
            continue
        key = normalise_title(title)
        if key:
            edition_groups[key].append(i)

    # For each group with multiple editions, find the index of the newest one
    # (highest edition number, then latest year as tiebreaker)
    older_edition_indices = set()
    for key, indices in edition_groups.items():
        if len(indices) < 2:
            continue
        # Sort by edition number desc, then year desc
        def sort_key(i):
            ed = get_edition_num(records[i])
            yr = get_year(records[i]) or 0
            # Primary: edition number, secondary: year — both descending
            # If no explicit edition marker, year alone determines newest
            return (ed if ed > 1 else 0, yr)
        sorted_indices = sorted(indices, key=sort_key, reverse=True)
        newest = sorted_indices[0]
        for older in sorted_indices[1:]:
            older_edition_indices.add(older)

    # Pre-compute volume groups: if any volume in a set is kept, keep all
    # Normalise title by stripping volume info, group by base title
    def normalise_volume_title(title):
        t = title.lower()
        # Remove volume markers: "volume I", "vol. 2", "deel 1", "band II", "tome 3"
        t = re.sub(r"[.,;]?\s*(volume|vol|deel|band|tome|part|partie|bd|fasc)\s*\.?\s*([ivxlcdm\d]+).*", "", t, flags=re.I)
        t = re.sub(r"[^a-z0-9 ]", " ", t)
        t = re.sub(r"\s+", " ", t).strip()
        return t[:60]

    volume_groups = defaultdict(list)
    for i, rec in enumerate(records):
        title = gf(rec, "T1", "TI")
        if not title:
            continue
        key = normalise_volume_title(title)
        base = normalise_title(title)
        # Only group as volumes if the normalised volume title differs from base title
        # (i.e. volume info was actually stripped)
        if key and key != base:
            volume_groups[key].append(i)

    # We'll resolve which volume groups to keep after the first pass
    # Store indices that belong to a multi-volume set
    volume_set_indices = {}  # index -> group key
    for key, indices in volume_groups.items():
        if len(indices) > 1:
            for i in indices:
                volume_set_indices[i] = key

    # Pre-compute translation groups: same author + year + Barnard, different language
    # Group key: first author (normalised) + year + barnard code
    def normalise_author(rec):
        a = gf(rec, "A1", "A2").lower()
        # Take first author surname only
        a = re.sub(r"[^a-z ]", "", a).strip()
        return " ".join(a.split()[:2])

    translation_groups = defaultdict(list)
    for i, rec in enumerate(records):
        author = normalise_author(rec)
        year = str(get_year(rec) or "")
        barnard = gf(rec, "U4").strip().upper()[:3]  # first 3 chars of Barnard
        lang = gf(rec, "U3").strip().lower()
        if author and year and barnard:
            key = f"{author}|{year}|{barnard}"
            translation_groups[key].append((i, lang))

    # Find groups with multiple languages — mark non-English as translation duplicates
    ENGLISH_VARIANTS = {"english", "eng", "en", "anglais", "engels"}
    translation_weed_indices = set()  # indices of non-English versions to weed
    for key, items in translation_groups.items():
        if len(items) < 2:
            continue
        langs = [lang for _, lang in items]
        has_english = any(l in ENGLISH_VARIANTS for l in langs)
        has_multiple_langs = len(set(langs)) > 1
        if has_english and has_multiple_langs:
            for i, lang in items:
                if lang not in ENGLISH_VARIANTS:
                    translation_weed_indices.add(i)

    # Process
    # Only process records from the first file; additional files are for stats only
    process_records = records[:primary_count]
    print(f"\nProcessing {len(process_records):,} records...")
    t0      = datetime.now()
    output_rows = []
    counts  = {"KEEP": 0, "WEED": 0, "REVIEW": 0, "SKIP": 0}

    for i, rec in enumerate(process_records):
        result = apply_rules(rec, records, borrowed, isbn_counts, older_edition=(i in older_edition_indices), translation_duplicate=(i in translation_weed_indices), barnard_counts=barnard_counts)
        rec_val = result["recommendation"]
        counts[rec_val] = counts.get(rec_val, 0) + 1

        output_rows.append({
            "row_index": i,
            "rec":       rec,
            "result":    result,
            "title":     gf(rec, "T1", "TI") or "Untitled",
            "author":    get_authors(rec),
            "year":      get_year(rec) or "",
            "rec_type":  gf(rec, "TY").upper(),
            "isbn":      get_isbn(rec),
            "location":  gf(rec, "U2").strip(),
        })

        if (i+1) % 500 == 0 or (i+1) == len(records):
            elapsed = (datetime.now() - t0).total_seconds()
            rate    = (i+1) / elapsed if elapsed > 0 else 0
            print(f"  {i+1:>6,} / {len(records):,}  "
                  f"KEEP={counts['KEEP']}  WEED={counts['WEED']}  "
                  f"REVIEW={counts['REVIEW']}  SKIP={counts['SKIP']}  "
                  f"({rate:.0f}/s)")

    # ── Second pass: upgrade WEED volumes to KEEP if another volume is kept ──
    kept_volume_groups = set()
    for row in output_rows:
        i = row["row_index"]
        if row["result"]["recommendation"] == "KEEP" and i in volume_set_indices:
            kept_volume_groups.add(volume_set_indices[i])

    if kept_volume_groups:
        upgraded = 0
        for row in output_rows:
            i = row["row_index"]
            if (row["result"]["recommendation"] == "WEED"
                    and i in volume_set_indices
                    and volume_set_indices[i] in kept_volume_groups):
                row["result"]["recommendation"] = "KEEP"
                row["result"]["reasoning"] += " | Kept: part of multi-volume set where another volume is kept."
                row["result"]["flags"].append({
                    "criterion": "Volume set",
                    "detail": "Part of multi-volume set — another volume is kept",
                    "severity": "keep"
                })
                upgraded += 1
        if upgraded:
            print(f"  Volume sets: {upgraded} volumes upgraded from WEED to KEEP")

    # Report translation duplicates
    trans_weeded = sum(1 for row in output_rows if any(f["criterion"] == "Translation" for f in row["result"]["flags"]))
    if trans_weeded:
        print(f"  Translations: {trans_weeded} non-English duplicates flagged")

    elapsed = (datetime.now() - t0).total_seconds()
    print(f"\nDone in {elapsed:.1f}s  ({len(process_records)/elapsed:.0f} records/sec)")
    print(f"  KEEP:   {counts['KEEP']:>6,}")
    print(f"  WEED:   {counts['WEED']:>6,}")
    print(f"  REVIEW: {counts['REVIEW']:>6,}")
    print(f"  SKIP:   {counts['SKIP']:>6,}")

    # Export
    print(f"\nExporting to {args.out}...")
    export_xlsx(output_rows, args.out)
    print(f"Done — {args.out}")


if __name__ == "__main__":
    main()
