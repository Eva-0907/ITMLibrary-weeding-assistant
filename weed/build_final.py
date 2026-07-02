#!/usr/bin/env python3
"""
build_final.py — ITG Antwerp Library Weeding Project
Merges the weeding report (weed9.py output) with UniCat lookup results
(unicat_lookup2.py output) into a single final Excel file.

Steps:
  1. Read weeding report (default: weeding_report9.xlsx)
  2. Read UniCat results (default: UniCat-lookup-results3.xlsx)
  3. Join on Bib# — upgrade any WEED sole-holder rows to REVIEW
  4. Write ITM_weeding_final.xlsx with two sheets:
       - Library Collection  (dark blue header)
       - Department Books    (purple header)

Usage:
    python3 build_final.py
    python3 build_final.py --weed weeding_report9.xlsx --unicat UniCat-lookup-results3.xlsx --out ITM_weeding_final.xlsx
"""

import sys
import argparse
from pathlib import Path

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter


FILL = {
    "KEEP":   PatternFill("solid", fgColor="DDFFDD"),
    "WEED":   PatternFill("solid", fgColor="FFDDDD"),
    "REVIEW": PatternFill("solid", fgColor="FFFFCC"),
    "SKIP":   PatternFill("solid", fgColor="EEEEEE"),
}
HEADER_BLUE   = PatternFill("solid", fgColor="1E3A5F")   # Library Collection
HEADER_PURPLE = PatternFill("solid", fgColor="4A235A")   # Department Books


def find_col(ws, name):
    """Return 1-based column index for a header name (case-insensitive), or None."""
    for cell in ws[1]:
        if cell.value and str(cell.value).strip().lower() == name.lower():
            return cell.column
    return None


def load_unicat_results(path):
    """
    Read UniCat-lookup-results3.xlsx.
    Returns dict: bib_number (str) -> {"result": str, "url": str}
    Falls back to ISBN key if Bib# not present.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    col_bib    = find_col(ws, "Bib#")
    col_isbn   = find_col(ws, "ISBN")
    col_result = find_col(ws, "UniCat Result")
    col_url    = find_col(ws, "UniCat URL")

    if not col_result:
        print("ERROR: 'UniCat Result' column not found in UniCat file.")
        sys.exit(1)

    results = {}
    for row in range(2, ws.max_row + 1):
        result = ws.cell(row=row, column=col_result).value
        url    = ws.cell(row=row, column=col_url).value if col_url else ""
        if not result:
            continue
        # Prefer Bib# as join key; fall back to ISBN
        key = None
        if col_bib:
            key = str(ws.cell(row=row, column=col_bib).value or "").strip()
        if not key and col_isbn:
            key = str(ws.cell(row=row, column=col_isbn).value or "").strip()
        if key:
            results[key] = {"result": str(result).strip(), "url": str(url or "").strip()}

    return results


def load_weeding_report(path):
    """
    Read all rows from all sheets of the weeding report.
    Returns (headers, rows_lib, rows_dept) where each row is a list of cell values.
    """
    wb = openpyxl.load_workbook(path, data_only=True)

    # Support single-sheet or two-sheet layout
    sheet_names = wb.sheetnames
    ws_lib  = wb["Library Collection"] if "Library Collection" in sheet_names else wb.active
    ws_dept = wb["Department Books"]   if "Department Books"   in sheet_names else None

    def read_sheet(ws):
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))
        return rows

    all_lib  = read_sheet(ws_lib)
    all_dept = read_sheet(ws_dept) if ws_dept else []

    headers = all_lib[0] if all_lib else []
    rows_lib  = all_lib[1:]
    rows_dept = all_dept[1:] if all_dept else []

    return headers, rows_lib, rows_dept


def merge(headers, rows, unicat_results):
    """
    Add UniCat Result and UniCat URL columns to rows.
    For WEED rows that are sole holders, upgrade to REVIEW and append to Reasoning.
    Returns (new_headers, new_rows, stats).
    """
    col_rec      = next((i for i, h in enumerate(headers) if h and str(h).lower() == "recommendation"), None)
    col_bib      = next((i for i, h in enumerate(headers) if h and str(h).lower() == "bib#"), None)
    col_isbn     = next((i for i, h in enumerate(headers) if h and str(h).lower() == "isbn"), None)
    col_reason   = next((i for i, h in enumerate(headers) if h and str(h).lower() == "reasoning"), None)

    if col_rec is None:
        print("ERROR: 'Recommendation' column not found in weeding report.")
        sys.exit(1)

    new_headers = list(headers) + ["UniCat Result", "UniCat URL"]
    upgraded = 0
    not_found = 0
    new_rows = []

    for row in rows:
        row = list(row)
        rec  = str(row[col_rec] or "").strip() if col_rec is not None else ""
        bib  = str(row[col_bib]  or "").strip() if col_bib  is not None else ""
        isbn = str(row[col_isbn] or "").strip() if col_isbn is not None else ""

        # Look up by Bib# first, then ISBN
        uc = unicat_results.get(bib) or unicat_results.get(isbn)

        if uc:
            uc_result = uc["result"]
            uc_url    = uc["url"]
            # Upgrade WEED sole holders to REVIEW
            if rec == "WEED" and uc_result == "Not held elsewhere":
                row[col_rec] = "REVIEW"
                upgraded += 1
                if col_reason is not None:
                    existing = str(row[col_reason] or "")
                    row[col_reason] = existing + " | REVIEW: no other Belgian library holds this — verify before weeding."
        else:
            uc_result = ""
            uc_url    = ""
            if rec == "WEED" and (bib or isbn):
                not_found += 1

        row.append(uc_result)
        row.append(uc_url)
        new_rows.append(row)

    stats = {"upgraded": upgraded, "not_found": not_found}
    return new_headers, new_rows, stats


def write_sheet(ws, headers, rows, header_fill):
    col_widths = {
        "title": 50, "author": 25, "year": 6, "type": 8, "bib#": 10,
        "isbn": 14, "barnard": 30, "retention flag": 8, "call number": 12,
        "language": 6, "location": 14, "recommendation": 14,
        "circulated": 10, "historically significant": 12,
        "historical reasons": 40, "triggered rules": 60,
        "check unicat": 12, "reasoning": 60,
        "unicat result": 18, "unicat url": 50,
    }
    fill_map = {"WEED": "FFDDDD", "KEEP": "DDFFDD", "REVIEW": "FFFFCC", "SKIP": "EEEEEE"}

    # Find recommendation column index (0-based in headers list)
    rec_idx = next((i for i, h in enumerate(headers) if h and str(h).lower() == "recommendation"), None)

    # Write header row
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = header_fill
        cell.alignment = Alignment(wrap_text=True)
        w = col_widths.get(str(h or "").lower(), 15)
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Write data rows
    for ri, row in enumerate(rows, 2):
        rec_val = str(row[rec_idx] or "").strip() if rec_idx is not None else ""
        row_fill = PatternFill("solid", fgColor=fill_map.get(rec_val, "FFFFFF"))
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill      = row_fill
            cell.alignment = Alignment(wrap_text=True)

    ws.freeze_panes = "A2"
    if ws.max_row > 1:
        ws.auto_filter.ref = ws.dimensions


def main():
    parser = argparse.ArgumentParser(description="Merge weeding report + UniCat results — ITG Antwerp")
    parser.add_argument("--weed",   default="weeding_report9.xlsx",      help="Weeding report (default: weeding_report9.xlsx)")
    parser.add_argument("--unicat", default="UniCat-lookup-results3.xlsx", help="UniCat results (default: UniCat-lookup-results3.xlsx)")
    parser.add_argument("--out",    default="ITM_weeding_final.xlsx",     help="Output file (default: ITM_weeding_final.xlsx)")
    args = parser.parse_args()

    # Validate inputs
    for p in (args.weed, args.unicat):
        if not Path(p).exists():
            print(f"ERROR: {p} not found.")
            sys.exit(1)

    print(f"Loading weeding report: {args.weed}")
    headers, rows_lib, rows_dept = load_weeding_report(args.weed)
    print(f"  Library Collection: {len(rows_lib):,} rows")
    print(f"  Department Books:   {len(rows_dept):,} rows")

    print(f"\nLoading UniCat results: {args.unicat}")
    unicat_results = load_unicat_results(args.unicat)
    print(f"  {len(unicat_results):,} UniCat results loaded")

    print("\nMerging Library Collection...")
    new_headers, new_rows_lib, stats_lib = merge(headers, rows_lib, unicat_results)
    print(f"  Sole-holder WEED→REVIEW upgrades: {stats_lib['upgraded']}")
    if stats_lib['not_found']:
        print(f"  WEED rows with no UniCat result:  {stats_lib['not_found']} (no ISBN/Bib# match)")

    print("Merging Department Books...")
    _, new_rows_dept, stats_dept = merge(headers, rows_dept, unicat_results)
    print(f"  Sole-holder WEED→REVIEW upgrades: {stats_dept['upgraded']}")

    # Tally final recommendations
    rec_idx = next((i for i, h in enumerate(new_headers) if h and str(h).lower() == "recommendation"), None)
    counts = {"KEEP": 0, "WEED": 0, "REVIEW": 0, "SKIP": 0}
    for row in new_rows_lib + new_rows_dept:
        r = str(row[rec_idx] or "").strip() if rec_idx is not None else ""
        if r in counts:
            counts[r] += 1

    print(f"\nFinal recommendation counts (all sheets):")
    for k, v in counts.items():
        print(f"  {k:<8} {v:>6,}")

    print(f"\nWriting {args.out}...")
    wb = Workbook()

    ws_lib = wb.active
    ws_lib.title = "Library Collection"
    write_sheet(ws_lib, new_headers, new_rows_lib, HEADER_BLUE)

    ws_dept = wb.create_sheet("Department Books")
    write_sheet(ws_dept, new_headers, new_rows_dept, HEADER_PURPLE)

    wb.save(args.out)
    print(f"Done — {args.out}")
    print(f"  Sheet 'Library Collection': {len(new_rows_lib):,} records")
    print(f"  Sheet 'Department Books':   {len(new_rows_dept):,} records")


if __name__ == "__main__":
    main()
